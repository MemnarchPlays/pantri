#!/usr/bin/env python3
"""Discord bot for the pantry tracker. Run: py discord_bot.py"""

import os
import json
from pathlib import Path

import discord
from dotenv import load_dotenv
from openpyxl import load_workbook
from pantry_utils import COLS, to_title, read_env, load_all_recipes

load_dotenv(override=True)

XLSX           = Path(__file__).parent / 'Food in Storage.xlsx'
DATA_DIR       = Path(__file__).parent / 'data'
EXCLUDE_SHEETS = {'Minimums'}


def load_minimums():
    wb = load_workbook(XLSX)
    if 'Minimums' not in wb.sheetnames:
        return {}
    ws = wb['Minimums']
    mins = {}
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=1).value
        qty  = ws.cell(row=r, column=2).value
        if item and qty is not None:
            mins[str(item).strip().lower()] = qty
    return mins


def set_minimum(item_title, qty):
    wb = load_workbook(XLSX)
    if 'Minimums' not in wb.sheetnames:
        ws = wb.create_sheet('Minimums')
        ws.cell(row=1, column=1).value = 'Item'
        ws.cell(row=1, column=2).value = 'Min Qty'
    else:
        ws = wb['Minimums']
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or '').lower() == item_title.lower():
            ws.cell(row=r, column=2).value = qty
            wb.save(XLSX)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = item_title
    ws.cell(row=row, column=2).value = qty
    wb.save(XLSX)


def delete_minimum(item_title):
    wb = load_workbook(XLSX)
    if 'Minimums' not in wb.sheetnames:
        return
    ws = wb['Minimums']
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or '').lower() == item_title.lower():
            ws.delete_rows(r)
            wb.save(XLSX)
            return


def load_location_sheets():
    wb = load_workbook(XLSX)
    return [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]


LOCATION_SHEETS = load_location_sheets()

intents = discord.Intents.default()
intents.message_content = True
intents.reactions = True
client = discord.Client(intents=intents)

# Tracks in-progress !add sessions per user: {user_id: {step, data}}
pending_adds = {}

# Tracks location-picker messages: {message_id: user_id}
location_pickers = {}

# Tracks recipe-picker messages: {message_id: {'uid': user_id, 'recipes': [...]}}
recipe_pickers = {}

NUMBER_EMOJIS = ['1️⃣','2️⃣','3️⃣','4️⃣','5️⃣','6️⃣','7️⃣','8️⃣','9️⃣','🔟']

KNOWN_UNITS = {
    'can', 'cans', 'bag', 'bags', 'box', 'boxes', 'jar', 'jars',
    'bottle', 'bottles', 'pack', 'packs', 'oz', 'lb', 'lbs',
    'pound', 'pounds', 'liter', 'liters', 'litre', 'litres',
    'gallon', 'gallons', 'cup', 'cups', 'tbsp', 'tsp',
    'piece', 'pieces', 'roll', 'rolls', 'loaf', 'loaves',
    'bunch', 'bunches', 'dozen', 'count', 'ct', 'pkg', 'g', 'kg',
}


def purge_empty_rows():
    wb = load_workbook(XLSX)
    changed = False
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        to_delete = [r for r in range(2, ws.max_row + 1) if not ws.cell(row=r, column=1).value]
        for r in reversed(to_delete):
            ws.delete_rows(r)
        if to_delete:
            changed = True
    if changed:
        wb.save(XLSX)


def get_all_rows():
    wb = load_workbook(XLSX)
    sheets = [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]
    rows = []
    for sheet_name in sheets:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
            item = vals[0]
            if any(v for v in vals) and item and not str(item).startswith('='):
                rows.append(dict(zip(COLS, vals)))
    return rows


def find_item_in(name, location):
    """Return (wb, ws, row_idx, data) for matching item within a specific sheet."""
    wb = load_workbook(XLSX)
    if location not in wb.sheetnames:
        return None
    ws = wb[location]
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        if any(vals) and str(vals[0] or '').lower() == name.lower():
            return wb, ws, r, dict(zip(COLS, vals))
    return None


def find_item(name):
    """Return (wb, ws, row_idx, data) for the first matching item, or None."""
    wb = load_workbook(XLSX)
    for sheet_name in LOCATION_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
            if any(vals) and str(vals[0] or '').lower() == name.lower():
                return wb, ws, r, dict(zip(COLS, vals))
    return None


def increment_quantity(name, delta, location=None):
    """Add delta to the item's quantity. If location given, only match within that sheet."""
    result = find_item(name) if not location else find_item_in(name, location)
    if not result:
        return None
    wb, ws, row_idx, data = result
    try:
        current = float(str(data.get('Quantity') or 0).strip())
    except ValueError:
        current = 0
    new_qty = current + delta
    new_qty = int(new_qty) if new_qty == int(new_qty) else new_qty
    ws.cell(row=row_idx, column=2).value = new_qty
    wb.save(XLSX)
    return new_qty, data.get('Unit') or ''


def set_quantity(name, qty, location=None):
    """Set item quantity to an exact value."""
    result = find_item(name) if not location else find_item_in(name, location)
    if not result:
        return None
    wb, ws, row_idx, data = result
    ws.cell(row=row_idx, column=2).value = qty
    wb.save(XLSX)
    return qty, data.get('Unit') or ''


async def maybe_alert_low_stock(item_name, new_qty, unit=''):
    """Send a low-stock alert to the configured channel if item dropped below minimum."""
    channel_id = read_env().get('DISCORD_ALERT_CHANNEL', '').strip()
    if not channel_id:
        return
    mins = load_minimums()
    min_qty = mins.get(item_name.lower())
    if min_qty is None:
        try:
            threshold = float(read_env().get('LOW_STOCK_THRESHOLD', '0'))
        except ValueError:
            threshold = 0
        if threshold <= 0:
            return
        min_qty = threshold
    if new_qty >= min_qty:
        return
    try:
        channel = client.get_channel(int(channel_id))
        if channel:
            u = f' {unit}' if unit else ''
            cur = int(new_qty) if new_qty == int(new_qty) else new_qty
            mn  = int(min_qty) if min_qty == int(min_qty) else min_qty
            await channel.send(f'⚠️ **{item_name}** is running low — **{cur}{u}** remaining (minimum: {mn}{u})')
    except Exception:
        pass


def fmt_item(r):
    qty  = str(r.get('Quantity') or '').strip()
    unit = str(r.get('Unit') or '').strip()
    loc  = str(r.get('Location') or '').strip()
    parts = []
    if qty or unit:
        parts.append(f"{qty} {unit}".strip())
    if loc:
        parts.append(loc)
    suffix = ' | '.join(p for p in parts if p)
    return f"**{r['Item']}**" + (f" — {suffix}" if suffix else '')


def cmd_stock(term):
    rows = get_all_rows()
    matches = [r for r in rows if term.lower() in str(r.get('Item', '')).lower()]
    if not matches:
        return f'Nothing matching **{term}** found in storage.'
    return '\n'.join(fmt_item(r) for r in matches)


def cmd_list(location_filter=None):
    purge_empty_rows()
    rows = get_all_rows()
    if location_filter:
        rows = [r for r in rows if location_filter.lower() in str(r.get('Location', '')).lower()]
    if not rows:
        return 'No items found.'
    header = f"**Pantry — {location_filter or 'All Locations'}** ({len(rows)} items)\n"
    return header + '\n'.join(f"• {fmt_item(r)}" for r in rows)


def cmd_add_save(data):
    sheet_name = data.get('Sheet', '')
    if sheet_name not in LOCATION_SHEETS:
        sheet_name = LOCATION_SHEETS[0]
    wb = load_workbook(XLSX)
    ws = wb[sheet_name]
    row = next((r for r in range(2, ws.max_row + 2) if not ws.cell(row=r, column=1).value), ws.max_row + 1)
    save_data = {
        'Item':     data.get('Item'),
        'Quantity': data.get('Quantity'),
        'Unit':     data.get('Unit'),
        'Location': sheet_name,
    }
    for col, field in enumerate(COLS, 1):
        ws.cell(row=row, column=col).value = save_data.get(field) or None
    wb.save(XLSX)


def item_exists(item_name, location=None):
    """If location given, only match within that sheet. Otherwise check all sheets."""
    wb = load_workbook(XLSX)
    all_sheets = [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]
    sheets = [location] if location and location in all_sheets else all_sheets
    for sheet_name in sheets:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=1).value
            if cell and str(cell).lower() == item_name.lower():
                return True
    return False


ADD_STEPS = [
    ('Item',     'What is the item name?',                    False),
    ('Quantity', 'How many / how much? (or type `skip`)',     True),
    ('Unit',     'What unit? Include size if relevant — e.g. `15oz can`, `28oz can`, `bag`, `1lb` (or `skip`)', True),
    ('Sheet',    'LOCATION_PICKER',                           False),
]


async def send_location_picker(channel, uid):
    lines = ['**Pick a location:**']
    for i, loc in enumerate(LOCATION_SHEETS):
        lines.append(f"{NUMBER_EMOJIS[i]} {loc}")
    msg = await channel.send('\n'.join(lines))
    for i in range(len(LOCATION_SHEETS)):
        await msg.add_reaction(NUMBER_EMOJIS[i])
    location_pickers[msg.id] = uid


def find_recipes(term):
    """Return matching recipes using substring then word-level fallback."""
    all_recipes = load_all_recipes()
    term_lower = term.lower()
    # Exact substring match first
    matches = [r for r in all_recipes if term_lower in r.get('name', '').lower()]
    if not matches:
        # Word-level: all words in term must appear somewhere in the name
        words = term_lower.split()
        matches = [r for r in all_recipes if all(w in r.get('name', '').lower() for w in words)]
    return matches


async def send_recipe_picker(channel, uid, matches):
    lines = ['**Which recipe?**']
    for i, r in enumerate(matches[:10]):
        lines.append(f"{NUMBER_EMOJIS[i]} {r['name']}")
    msg = await channel.send('\n'.join(lines))
    for i in range(min(len(matches), 10)):
        await msg.add_reaction(NUMBER_EMOJIS[i])
    recipe_pickers[msg.id] = {'uid': uid, 'recipes': matches[:10]}


def build_recipe_chunks(r):
    macros = r.get('macros_per_serving', {})
    have   = {str(row.get('Item', '')).lower() for row in get_all_rows() if row.get('Item')}
    lines  = [
        f"**{r['name']}** — {r.get('meal_type', '')} | Serves {r.get('servings', '?')} | Prep {r.get('prep_time', '?')} | Cook {r.get('cook_time', '?')}",
        f"*Calories: {macros.get('calories','?')} | Protein: {macros.get('protein_g','?')}g | Carbs: {macros.get('carbs_g','?')}g | Fat: {macros.get('fat_g','?')}g*",
        '', '**Ingredients**',
    ]
    for ing in r.get('ingredients', []):
        name = ing.get('item', '')
        icon = '✅' if any(name.lower() in h or h in name.lower() for h in have) else '•'
        lines.append(f"{icon} {ing.get('amount', '')} {name}".strip())
    lines += ['', '**Instructions**']
    for i, step in enumerate(r.get('instructions', []), 1):
        lines.append(f"{i}. {step}")
    subs = r.get('substitutions', {})
    if subs:
        lines += ['', '**Substitutions**']
        for key, val in subs.items():
            lines.append(f"*{key.replace('_', ' ').title()}:* {val}")
    if r.get('notes'):
        lines += ['', f"**Notes:** {r['notes']}"]
    chunks, current = [], ''
    for line in lines:
        if len(current) + len(line) + 1 > 1900:
            chunks.append(current)
            current = line + '\n'
        else:
            current += line + '\n'
    if current:
        chunks.append(current)
    return chunks


def cmd_recipe(term):
    matches = find_recipes(term)
    if not matches:
        return [f'No recipe found matching **{term}**.'], None
    if len(matches) > 1:
        return None, matches

    r = matches[0]
    return build_recipe_chunks(r), None


MEAL_TYPES = {'breakfast', 'lunch', 'dinner', 'snack', 'dessert'}

def cmd_can_make(filter_term=None):
    rows = get_all_rows()
    have = {str(r.get('Item', '')).lower() for r in rows if r.get('Item')}

    results = []
    for recipe in load_all_recipes():
        # Apply filter
        if filter_term:
            meal = recipe.get('meal_type', '').lower()
            name = recipe.get('name', '').lower()
            if filter_term.lower() in MEAL_TYPES:
                if filter_term.lower() not in meal:
                    continue
            else:
                if filter_term.lower() not in name:
                    continue

        ingredients = recipe.get('ingredients', [])
        needed = [(i.get('item', '') if isinstance(i, dict) else str(i)).lower() for i in ingredients]
        if not needed:
            continue
        have_count = sum(1 for n in needed if any(n in h or h in n for h in have))
        total = len(needed)
        pct   = int(have_count / total * 100) if total else 0
        results.append((pct, have_count, total, recipe.get('name', 'Untitled')))

    results.sort(key=lambda x: x[0], reverse=True)

    filter_label = f' — {filter_term.title()}' if filter_term else ''
    lines = [f'**Can I Make This?{filter_label}**\n']

    shown = results[:25]
    for pct, have_count, total, name in shown:
        if pct == 100:
            icon = '✅'
        elif pct >= 60:
            icon = '🟡'
        elif pct > 0:
            icon = '🟠'
        else:
            icon = '❌'
        lines.append(f"{icon} **{name}** — {pct}% ({have_count}/{total} ingredients)")

    remaining = len(results) - len(shown)
    if remaining > 0:
        lines.append(f'\n_...and {remaining} more. Try `!canmake breakfast`, `!canmake lunch`, `!canmake dinner`, or `!canmake <name>` to filter._')
    elif not results:
        lines.append('No recipes found.')
    elif not have:
        lines.append('\n_Add items to your pantry with `!add` to see matches improve._')

    return '\n'.join(lines)


HELP_TEXT = """**Pantri Bot Commands**

**Recipes**
`!recipe <name>` — show a recipe (partial name ok)
`!canmake [meal or name]` — recipes you can make; filter by meal type or keyword

**Inventory**
`!add <item> [qty] [unit]` — add an item (bot asks for missing fields) · `!cancel` to abort
`!stock <item>` — check quantity · `!list [location]` — list all items
`!set <item> <qty>` — set exact quantity
`!remove <item> <qty>` — subtract quantity from an item (e.g. `!remove black beans 2`)
`!restock` — items below minimum · `!setmin <item> <qty>` — set a minimum

**Locations**
`!locations` — list · `!addlocation <name>` — add new

`!help` — this message"""


@client.event
async def on_ready():
    print(f'Pantry Bot online as {client.user}')


@client.event
async def on_message(message):
    if message.author == client.user:
        return

    content = message.content.strip()
    uid     = message.author.id

    # Handle in-progress !add wizard
    if uid in pending_adds and not content.startswith('!'):
        session = pending_adds[uid]
        step    = session['step']
        field, _, optional = ADD_STEPS[step]

        # If waiting for location, accept typed name or remind to use reactions
        if field == 'Sheet':
            match = next((l for l in LOCATION_SHEETS if l.lower() == content.lower()), None)
            if not match:
                locs = ', '.join(f'`{l}`' for l in LOCATION_SHEETS)
                await message.channel.send(f"Please pick a location using the reactions, or type one exactly: {locs}")
                return
            content = match

        raw   = content if not (optional and content.lower() == 'skip') else None
        value = to_title(raw) if raw and field == 'Item' else raw
        session['data'][field] = value

        # Advance to next unfilled step
        next_step = next(
            (i for i in range(session['step'] + 1, len(ADD_STEPS))
             if ADD_STEPS[i][0] not in session['data'] or session['data'][ADD_STEPS[i][0]] is None),
            len(ADD_STEPS)
        )
        session['step'] = next_step

        if session['step'] >= len(ADD_STEPS):
            data = session['data']
            if item_exists(data['Item']):
                del pending_adds[uid]
                delta = float(data['Quantity']) if data.get('Quantity') else 1
                result = increment_quantity(data['Item'], delta)
                if result:
                    new_qty, unit = result
                    await message.channel.send(f"✅ Updated **{data['Item']}** — now **{new_qty} {unit}**.".strip())
                    await maybe_alert_low_stock(data['Item'], new_qty, unit)
            else:
                cmd_add_save(data)
                del pending_adds[uid]
                await message.channel.send(f"✅ Added **{data['Item']}** to **{data.get('Sheet', 'storage')}**.")
        else:
            next_field, next_prompt, _ = ADD_STEPS[session['step']]
            if next_prompt == 'LOCATION_PICKER':
                await send_location_picker(message.channel, uid)
            else:
                await message.channel.send(next_prompt)
        return

    if content.startswith('!recipe '):
        term = content[8:].strip()
        chunks, matches = cmd_recipe(term)
        if matches:
            await send_recipe_picker(message.channel, uid, matches)
        else:
            for chunk in chunks:
                await message.channel.send(chunk)

    elif content.startswith('!stock '):
        term = content[7:].strip()
        await message.channel.send(cmd_stock(term))

    elif content.startswith('!add'):
        # Cancel any existing session for this user
        if uid in pending_adds:
            del pending_adds[uid]
            stale = [mid for mid, u in list(location_pickers.items()) if u == uid]
            for mid in stale:
                del location_pickers[mid]

        raw = content[4:].strip()
        if not raw:
            await message.channel.send('Usage: `!add <item> <quantity> <location>`')
        else:
            # Extract location — match longest location name first to handle multi-word names
            sheet = None
            raw_lower = raw.lower()
            for loc in sorted(LOCATION_SHEETS, key=len, reverse=True):
                if loc.lower() in raw_lower:
                    sheet = loc
                    raw = raw_lower.replace(loc.lower(), '').strip()
                    break
            else:
                raw = raw_lower

            # Find the numeric token — item is before it, unit is after it
            parts = raw.split()
            num_idx = next((i for i, p in enumerate(parts) if p.replace('.', '', 1).isdigit()), None)

            if num_idx is None:
                # No number found — everything is the item name
                qty  = None
                unit = None
                item = to_title(' '.join(parts))
            elif num_idx == 0:
                # Number first — try <qty> <unit> <item> or <qty> <item>
                qty = parts[0]
                rest = parts[1:]
                if rest and rest[0].lower() in KNOWN_UNITS:
                    unit = rest[0]
                    item = to_title(' '.join(rest[1:])) if rest[1:] else None
                else:
                    unit = None
                    item = to_title(' '.join(rest)) if rest else None
                if not item:
                    await message.channel.send('Please include an item name: e.g. `!add sprite 2 liters`')
                    return
            else:
                item = to_title(' '.join(parts[:num_idx]))
                qty  = parts[num_idx]
                unit = ' '.join(parts[num_idx + 1:]) or None

            # Always show location picker — existence check happens after location is chosen
            data = {'Item': item, 'Quantity': qty, 'Unit': unit}
            first_missing = next(
                (i for i, (field, _, _) in enumerate(ADD_STEPS)
                 if field not in data or data[field] is None), None
            )
            if first_missing is None:
                sheet_step = next(i for i, (f, _, _) in enumerate(ADD_STEPS) if f == 'Sheet')
                pending_adds[uid] = {'step': sheet_step, 'data': data}
                await send_location_picker(message.channel, uid)
            else:
                pending_adds[uid] = {'step': first_missing, 'data': data}
                _, prompt, _ = ADD_STEPS[first_missing]
                if prompt == 'LOCATION_PICKER':
                    await send_location_picker(message.channel, uid)
                else:
                    await message.channel.send(prompt)

    elif content == '!list':
        reply = cmd_list()
        for chunk in [reply[i:i+1900] for i in range(0, len(reply), 1900)]:
            await message.channel.send(chunk)

    elif content.startswith('!list '):
        loc = content[6:].strip()
        await message.channel.send(cmd_list(loc))

    elif content == '!canmake' or content.startswith('!canmake '):
        filter_term = content[9:].strip() or None
        reply = cmd_can_make(filter_term)
        for chunk in [reply[i:i+1900] for i in range(0, len(reply), 1900)]:
            await message.channel.send(chunk)

    elif content.startswith('!addlocation '):
        new_loc = to_title(content[13:].strip())
        if not new_loc:
            await message.channel.send('Usage: `!addlocation <name>`')
        elif new_loc in LOCATION_SHEETS:
            await message.channel.send(f'**{new_loc}** already exists.')
        else:
            wb = load_workbook(XLSX)
            ws = wb.create_sheet(new_loc)
            for col, header in enumerate(COLS, 1):
                ws.cell(row=1, column=col).value = header
            wb.save(XLSX)
            LOCATION_SHEETS.append(new_loc)
            await message.channel.send(f'✅ Added location **{new_loc}**.')

    elif content == '!locations':
        await message.channel.send('**Storage Locations:**\n' + '\n'.join(f'• {l}' for l in LOCATION_SHEETS))

    elif content == '!restock':
        rows = get_all_rows()
        mins = load_minimums()

        # Build stock totals keyed by lowercase item name
        stock = {}
        units = {}
        for r in rows:
            name = str(r.get('Item') or '').strip().lower()
            if not name:
                continue
            try:
                qty = float(str(r.get('Quantity') or 0))
            except ValueError:
                qty = 0
            stock[name] = stock.get(name, 0) + qty
            if name not in units:
                units[name] = str(r.get('Unit') or '').strip()

        # Check every minimum — items not in pantry count as 0
        low = []
        for item_lower, min_qty in sorted(mins.items()):
            current = stock.get(item_lower, 0)
            if current < min_qty:
                low.append((to_title(item_lower), current, min_qty, units.get(item_lower, '')))

        if not low:
            await message.channel.send('✅ Everything is stocked up — nothing below minimum.')
        else:
            await message.channel.send(f'🛒 **{len(low)} item(s) need restocking:**')
            for item, current, min_qty, unit in low:
                u = f' {unit}' if unit else ''
                await message.channel.send(f'• **{item}** — {int(current) if current == int(current) else current}{u} (min: {int(min_qty) if min_qty == int(min_qty) else min_qty}{u})')

    elif content.startswith('!setmin '):
        parts = content[8:].strip().split()
        num_idx = next((i for i, p in enumerate(parts) if p.replace('.', '', 1).isdigit()), None)
        if num_idx is None or num_idx == 0:
            await message.channel.send('Usage: `!setmin <item> <minimum quantity>` — e.g. `!setmin black beans 3`')
        else:
            item = to_title(' '.join(parts[:num_idx]))
            min_qty = float(parts[num_idx])
            min_qty = int(min_qty) if min_qty == int(min_qty) else min_qty
            set_minimum(item, min_qty)
            await message.channel.send(f'✅ Minimum for **{item}** set to **{min_qty}**.')

    elif content.startswith('!set '):
        parts = content[5:].strip().split()
        num_idx = next((i for i, p in enumerate(parts) if p.replace('.','',1).isdigit()), None)
        if num_idx is None or num_idx == 0:
            await message.channel.send('Usage: `!set <item> <quantity>` — e.g. `!set black beans 5`')
        else:
            item = to_title(' '.join(parts[:num_idx]))
            qty  = float(parts[num_idx])
            qty  = int(qty) if qty == int(qty) else qty
            result = set_quantity(item, qty)
            if result:
                new_qty, unit = result
                await message.channel.send(f'✅ **{item}** set to **{new_qty}{f" {unit}" if unit else ""}**.')
                await maybe_alert_low_stock(item, new_qty, unit)
            else:
                await message.channel.send(f'❌ **{item}** not found in storage.')

    elif content.startswith('!remove '):
        parts = content[8:].strip().split()
        num_idx = next((i for i, p in enumerate(parts) if p.replace('.','',1).isdigit()), None)
        if num_idx is None or num_idx == 0:
            await message.channel.send('Usage: `!remove <item> <quantity>` — e.g. `!remove black beans 2`')
        else:
            item = to_title(' '.join(parts[:num_idx]))
            qty  = float(parts[num_idx])
            qty  = int(qty) if qty == int(qty) else qty
            result = increment_quantity(item, -qty)
            if result:
                new_qty, unit = result
                if new_qty < 0:
                    new_qty = 0
                    set_quantity(item, 0)
                await message.channel.send(f'✅ Removed **{qty}** from **{item}** — now **{new_qty}{f" {unit}" if unit else ""}**.')
                await maybe_alert_low_stock(item, new_qty, unit)
            else:
                await message.channel.send(f'❌ **{item}** not found in storage.')

    elif content == '!cancel':
        if uid in pending_adds:
            del pending_adds[uid]
            stale = [mid for mid, u in list(location_pickers.items()) if u == uid]
            for mid in stale:
                del location_pickers[mid]
            await message.channel.send('❌ Add cancelled.')
        else:
            await message.channel.send('Nothing to cancel.')

    elif content == '!help':
        await message.channel.send(HELP_TEXT)


@client.event
async def on_reaction_add(reaction, user):
    if user == client.user:
        return
    msg_id = reaction.message.id

    if msg_id in recipe_pickers:
        picker = recipe_pickers[msg_id]
        if user.id != picker['uid']:
            return
        emoji = str(reaction.emoji)
        if emoji not in NUMBER_EMOJIS:
            return
        idx = NUMBER_EMOJIS.index(emoji)
        if idx >= len(picker['recipes']):
            return
        del recipe_pickers[msg_id]
        await reaction.message.delete()
        for chunk in build_recipe_chunks(picker['recipes'][idx]):
            await reaction.message.channel.send(chunk)
        return

    if msg_id not in location_pickers:
        return
    uid = location_pickers[msg_id]
    if user.id != uid:
        return

    emoji = str(reaction.emoji)
    if emoji not in NUMBER_EMOJIS:
        return

    idx = NUMBER_EMOJIS.index(emoji)
    if idx >= len(LOCATION_SHEETS):
        return

    location = LOCATION_SHEETS[idx]
    del location_pickers[msg_id]
    await reaction.message.delete()

    if uid not in pending_adds:
        return

    session = pending_adds[uid]
    session['data']['Sheet'] = location

    # Advance to next unfilled step
    next_step = next(
        (i for i in range(session['step'] + 1, len(ADD_STEPS))
         if ADD_STEPS[i][0] not in session['data'] or session['data'][ADD_STEPS[i][0]] is None),
        len(ADD_STEPS)
    )
    session['step'] = next_step

    if session['step'] >= len(ADD_STEPS):
        data = session['data']
        if item_exists(data['Item'], location=data.get('Sheet')):
            del pending_adds[uid]
            delta = float(data['Quantity']) if data.get('Quantity') else 1
            result = increment_quantity(data['Item'], delta, location=data.get('Sheet'))
            if result:
                new_qty, unit = result
                await reaction.message.channel.send(f"✅ Updated **{data['Item']}** in **{location}** — now **{new_qty} {unit}**.".strip())
                await maybe_alert_low_stock(data['Item'], new_qty, unit)
        else:
            cmd_add_save(data)
            del pending_adds[uid]
            await reaction.message.channel.send(f"✅ Added **{data['Item']}** to **{location}**.")
    else:
        next_field, next_prompt, _ = ADD_STEPS[session['step']]
        if next_prompt == 'LOCATION_PICKER':
            await send_location_picker(reaction.message.channel, uid)
        else:
            await reaction.message.channel.send(next_prompt)


client.run(os.getenv('DISCORD_TOKEN'))
