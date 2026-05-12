"""Workbook helpers — read/write the pantry Excel file."""

from openpyxl import load_workbook
from pantry_utils import COLS, XLSX, EXCLUDE_SHEETS


def init_wb():
    """Create a blank workbook with just the Minimums sheet if the file doesn't exist."""
    if not XLSX.exists():
        from openpyxl import Workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        ws = wb.create_sheet('Minimums')
        ws.cell(row=1, column=1).value = 'Item'
        ws.cell(row=1, column=2).value = 'Min Qty'
        wb.save(XLSX)


def get_location_sheets():
    init_wb()
    wb = load_workbook(XLSX)
    return [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]


def load_wb():
    init_wb()
    return load_workbook(XLSX)


def save_wb(wb):
    from pantri.backup import backup_wb  # lazy import to avoid circular dependency
    dedup_workbook(wb)
    wb.save(XLSX)
    try:
        backup_wb()
    except Exception:
        pass


def dedup_workbook(wb):
    """Merge rows with the same item name AND unit within each sheet. Returns True if changes were made."""
    changed = False
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        seen = {}  # (item_lower, unit_lower) -> first row_idx
        to_clear = []
        for r in range(2, ws.max_row + 1):
            item = ws.cell(row=r, column=1).value
            if not item or str(item).startswith('='):
                continue
            unit = str(ws.cell(row=r, column=3).value or '').strip().lower()
            key = (str(item).strip().lower(), unit)
            if key in seen:
                first = seen[key]
                try:
                    q1 = float(str(ws.cell(row=first, column=2).value or 0))
                    q2 = float(str(ws.cell(row=r,     column=2).value or 0))
                    merged = q1 + q2
                    ws.cell(row=first, column=2).value = int(merged) if merged == int(merged) else merged
                except (ValueError, TypeError):
                    pass
                to_clear.append(r)
                changed = True
            else:
                seen[key] = r
        for r in reversed(to_clear):
            ws.delete_rows(r)
    return changed


def get_all_rows(wb):
    rows = []
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
            item = vals[0]
            if any(v for v in vals) and item and not str(item).startswith('='):
                rows.append((sheet_name, r, dict(zip(COLS, vals))))
    return rows


def get_location_info():
    wb = load_wb()
    result = []
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value)
        result.append((sheet_name, count))
    return result


def get_minimums():
    wb = load_wb()
    if 'Minimums' not in wb.sheetnames:
        return []
    ws = wb['Minimums']
    result = []
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=1).value
        qty  = ws.cell(row=r, column=2).value
        item_type = str(ws.cell(row=r, column=3).value or '').strip()
        if item:
            result.append({'item': str(item), 'qty': qty, 'type': item_type})
    return sorted(result, key=lambda x: (x['type'].lower() or 'zzz', x['item'].lower()))


def set_minimum(item_title, qty, item_type=''):
    wb = load_wb()
    if 'Minimums' not in wb.sheetnames:
        ws = wb.create_sheet('Minimums')
        ws.cell(row=1, column=1).value = 'Item'
        ws.cell(row=1, column=2).value = 'Min Qty'
        ws.cell(row=1, column=3).value = 'Type'
    else:
        ws = wb['Minimums']
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or '').lower() == item_title.lower():
            ws.cell(row=r, column=2).value = qty
            ws.cell(row=r, column=3).value = item_type or None
            save_wb(wb)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = item_title
    ws.cell(row=row, column=2).value = qty
    ws.cell(row=row, column=3).value = item_type or None
    save_wb(wb)


def delete_minimum(item_title):
    wb = load_wb()
    if 'Minimums' not in wb.sheetnames:
        return
    ws = wb['Minimums']
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or '').lower() == item_title.lower():
            ws.delete_rows(r)
            save_wb(wb)
            return
