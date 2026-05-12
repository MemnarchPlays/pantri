"""JSON state helpers — shopping list, exclusions, units, types."""

import json
import os
from pathlib import Path
from pantry_utils import XLSX, EXCLUDE_SHEETS

_data     = Path(os.environ['PANTRI_DATA']) if 'PANTRI_DATA' in os.environ else Path(__file__).parent.parent
STATE_DIR = _data / 'state'
STATE_DIR.mkdir(parents=True, exist_ok=True)
SHOPPING_LIST_FILE = STATE_DIR / 'shopping_list.json'
EXCLUSIONS_FILE    = STATE_DIR / 'exclusions.json'
UNITS_FILE         = STATE_DIR / 'units.json'
TYPES_FILE         = STATE_DIR / 'types.json'

DEFAULT_TYPES = [
    'Baking', 'Beverages', 'Canned Goods', 'Cleaning', 'Condiments',
    'Dairy', 'Dry Goods', 'Frozen', 'Oils', 'Produce', 'Spices',
]


def load_shopping_list():
    if not SHOPPING_LIST_FILE.exists():
        return []
    try:
        return json.loads(SHOPPING_LIST_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def save_shopping_list(items):
    SHOPPING_LIST_FILE.write_text(json.dumps(items, indent=2), encoding='utf-8')


def load_exclusions():
    if not EXCLUSIONS_FILE.exists():
        return []
    try:
        return json.loads(EXCLUSIONS_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def save_exclusions(items):
    EXCLUSIONS_FILE.write_text(json.dumps(items, indent=2), encoding='utf-8')


def load_units():
    if not UNITS_FILE.exists():
        return []
    try:
        return json.loads(UNITS_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def save_units(units):
    UNITS_FILE.write_text(json.dumps(sorted(units, key=str.lower), indent=2), encoding='utf-8')


def get_unit_info():
    from openpyxl import load_workbook
    units = load_units()
    if not XLSX.exists():
        return [(u, 0) for u in units]
    wb = load_workbook(XLSX)
    counts = {u.lower(): 0 for u in units}
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=3).value
            if cell_val:
                key = str(cell_val).strip().lower()
                if key in counts:
                    counts[key] += 1
    return [(u, counts[u.lower()]) for u in units]


def load_types():
    if not TYPES_FILE.exists():
        return list(DEFAULT_TYPES)
    try:
        data = json.loads(TYPES_FILE.read_text(encoding='utf-8'))
        return data if data else list(DEFAULT_TYPES)
    except Exception:
        return list(DEFAULT_TYPES)


def save_types(types):
    TYPES_FILE.write_text(json.dumps(types, indent=2), encoding='utf-8')


def is_excluded(item_name):
    name = item_name.strip().lower()
    for ex in load_exclusions():
        ex_name = ex.get('name', '').lower()
        if name == ex_name or name in ex_name or ex_name in name:
            return True
    return False
