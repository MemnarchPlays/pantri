#!/usr/bin/env python3
"""CLI pantry tracker — reads and writes Food in Storage.xlsx."""

import json
import sys
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

XLSX = Path(__file__).parent / 'Food in Storage.xlsx'
COLS = ['Item', 'Quantity', 'Unit', 'Location', 'Section', 'Slot', 'Expiration', 'Notes']
DATA_DIR = Path(__file__).parent / 'data'
LOCATION_SHEETS = ['Brown Cabinet', 'Pantry', 'End Hall Closet', 'Laundry Room', 'Kitchen']

HELP = """
Pantry CLI — commands:
  list                    Show all pantry items
  list <location>         Filter by location (e.g. "Brown Cabinet")
  search <term>           Search items by name
  add                     Add a new item (interactive)
  update <item name>      Update quantity/details of an item
  remove <item name>      Remove an item
  can-make                Show which recipes you can make right now
  help                    Show this message
"""


def load_wb():
    if not XLSX.exists():
        print(f'Error: {XLSX} not found.')
        sys.exit(1)
    return load_workbook(XLSX)


def save_wb(wb):
    wb.save(XLSX)


def get_all_rows(wb):
    """Read rows from all location sheets, tagging each with (sheet_name, row_index)."""
    rows = []
    for sheet_name in LOCATION_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            if any(v for v in row_vals):
                d = dict(zip(COLS, row_vals))
                rows.append((sheet_name, r, d))
    return rows


def find_first_empty(ws):
    for r in range(2, ws.max_row + 2):
        if not ws.cell(row=r, column=1).value:
            return r
    return ws.max_row + 1


def cmd_list(args):
    wb = load_wb()
    all_rows = get_all_rows(wb)
    filter_loc = ' '.join(args).strip().lower() if args else None

    if filter_loc:
        all_rows = [(s, r, d) for s, r, d in all_rows if filter_loc in str(d.get('Location', '')).lower()]

    if not all_rows:
        print('No items found.')
        return

    print(f'\n{"Item":<30} {"Qty":<8} {"Unit":<12} {"Location":<20} {"Section":<18} {"Slot":<12} {"Exp":<12} Notes')
    print('-' * 120)
    for _, _, d in all_rows:
        print(f'{str(d["Item"]):<30} {str(d["Quantity"] or ""):<8} {str(d["Unit"] or ""):<12} '
              f'{str(d["Location"] or ""):<20} {str(d["Section"] or ""):<18} '
              f'{str(d["Slot"] or ""):<12} {str(d["Expiration"] or ""):<12} {str(d["Notes"] or "")}')
    print(f'\n{len(all_rows)} item(s) shown.\n')


def cmd_search(args):
    if not args:
        print('Usage: search <term>')
        return
    term = ' '.join(args).lower()
    wb   = load_wb()
    rows = [(s, r, d) for s, r, d in get_all_rows(wb) if term in str(d.get('Item', '')).lower()]
    if not rows:
        print(f'No items matching "{term}".')
        return
    for _, _, d in rows:
        print(f'  {d["Item"]}  |  qty: {d["Quantity"]} {d["Unit"]}  |  {d["Location"]} / {d["Section"]} / {d["Slot"]}  |  exp: {d["Expiration"]}')


def cmd_add(_):
    print('\nAdd new pantry item (press Enter to skip optional fields):')
    item       = input('Item name: ').strip()
    if not item:
        print('Cancelled.')
        return
    quantity   = input('Quantity: ').strip()
    unit       = input('Unit (e.g. can, bag, oz): ').strip()
    print(f'Locations: {", ".join(LOCATION_SHEETS)}')
    location   = input('Location: ').strip()
    section    = input('Section (e.g. Top Shelf): ').strip()
    slot       = input('Slot (e.g. Left): ').strip()
    expiration = input('Expiration date (optional): ').strip()
    notes      = input('Notes (optional): ').strip()

    sheet_name = location if location in LOCATION_SHEETS else LOCATION_SHEETS[0]
    wb  = load_wb()
    ws  = wb[sheet_name]
    row = find_first_empty(ws)
    ws.cell(row=row, column=1).value = item
    ws.cell(row=row, column=2).value = quantity or None
    ws.cell(row=row, column=3).value = unit or None
    ws.cell(row=row, column=4).value = location or sheet_name
    ws.cell(row=row, column=5).value = section or None
    ws.cell(row=row, column=6).value = slot or None
    ws.cell(row=row, column=7).value = expiration or None
    ws.cell(row=row, column=8).value = notes or None
    save_wb(wb)
    print(f'Added: {item} → {sheet_name}')


def cmd_update(args):
    if not args:
        print('Usage: update <item name>')
        return
    name = ' '.join(args).lower()
    wb   = load_wb()
    all_rows = get_all_rows(wb)

    match = None
    for sheet_name, row_idx, d in all_rows:
        if name in str(d.get('Item', '')).lower():
            match = (sheet_name, row_idx, d)
            print(f'Found: {d["Item"]} in {sheet_name}')
            break

    if not match:
        print(f'Item not found: {name}')
        return

    sheet_name, row_idx, _ = match
    ws = wb[sheet_name]
    print('Enter new values (press Enter to keep current):')
    for col, field in enumerate(COLS, 1):
        current = ws.cell(row=row_idx, column=col).value
        new_val = input(f'  {field} [{current}]: ').strip()
        if new_val:
            ws.cell(row=row_idx, column=col).value = new_val

    save_wb(wb)
    print('Updated.')


def cmd_remove(args):
    if not args:
        print('Usage: remove <item name>')
        return
    name = ' '.join(args).lower()
    wb   = load_wb()

    for sheet_name, row_idx, d in get_all_rows(wb):
        cell_val = d.get('Item', '')
        if name in str(cell_val).lower():
            confirm = input(f'Remove "{cell_val}" from {sheet_name}? (y/n): ').strip().lower()
            if confirm == 'y':
                ws = wb[sheet_name]
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).value = None
                save_wb(wb)
                print(f'Removed: {cell_val}')
            else:
                print('Cancelled.')
            return

    print(f'Item not found: {name}')


def cmd_can_make(_):
    if not DATA_DIR.exists():
        print('No recipe data found.')
        return

    wb   = load_wb()
    have = {str(d.get('Item', '')).lower() for _, _, d in get_all_rows(wb) if d.get('Item')}

    print('\nRecipe availability:')
    found_any = False
    for path in sorted(DATA_DIR.glob('*.json')):
        with open(path, encoding='utf-8') as f:
            recipe = json.load(f)

        ingredients = recipe.get('ingredients', [])
        needed = []
        for ing in ingredients:
            item = ing.get('item', '') if isinstance(ing, dict) else str(ing)
            needed.append(item.lower())

        if not needed:
            continue

        have_count = sum(1 for n in needed if any(n in h or h in n for h in have))
        total      = len(needed)
        pct        = int(have_count / total * 100) if total else 0
        status     = 'YES' if pct == 100 else f'{pct}% ({have_count}/{total})'
        missing    = [n for n in needed if not any(n in h or h in n for h in have)]

        print(f'  {recipe.get("name", "Untitled"):35}  {status}')
        if missing and pct < 100:
            print(f'    Missing: {", ".join(missing[:5])}{"..." if len(missing) > 5 else ""}')
        found_any = True

    if not found_any:
        print('  No recipes with ingredients found.')
    print()


COMMANDS = {
    'list':     cmd_list,
    'search':   cmd_search,
    'add':      cmd_add,
    'update':   cmd_update,
    'remove':   cmd_remove,
    'can-make': cmd_can_make,
    'help':     lambda _: print(HELP),
}


def main():
    if len(sys.argv) < 2:
        print(HELP)
        return

    cmd  = sys.argv[1].lower()
    args = sys.argv[2:]

    if cmd in COMMANDS:
        COMMANDS[cmd](args)
    else:
        print(f'Unknown command: {cmd}')
        print(HELP)


if __name__ == '__main__':
    main()
