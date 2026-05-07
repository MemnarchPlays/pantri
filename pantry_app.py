#!/usr/bin/env python3
"""Flask web app for the pantry tracker. Run this file then open http://localhost:5000"""

import colorsys
import io
import json
import zipfile
import re
import shutil
import socket
import subprocess
import sys
import threading
import webbrowser
from datetime import datetime
from pathlib import Path
from flask import Flask, request, redirect, url_for, render_template_string, jsonify, send_file
from openpyxl import load_workbook
from pantry_utils import COLS, to_title, read_env

app = Flask(__name__)


def compute_theme(hex_color):
    try:
        h = hex_color.lstrip('#')
        r, g, b = int(h[0:2], 16)/255, int(h[2:4], 16)/255, int(h[4:6], 16)/255
    except Exception:
        r, g, b = 0.42, 0.18, 0.55
    hue, lum, sat = colorsys.rgb_to_hls(r, g, b)
    def to_hex(r, g, b):
        return '#{:02X}{:02X}{:02X}'.format(int(r*255), int(g*255), int(b*255))
    dark  = colorsys.hls_to_rgb(hue, max(0, lum * 0.70), sat)
    light = colorsys.hls_to_rgb(hue, min(1, lum + (1 - lum) * 0.75), sat * 0.4)
    return hex_color, to_hex(*dark), to_hex(*light)


@app.context_processor
def inject_theme():
    color = read_env().get('ACCENT_COLOR', '#6B2D8B')
    main, dark, light = compute_theme(color)
    css = f':root {{ --purple: {main}; --dark-purple: {dark}; --light-purple: {light}; }}'
    return {'theme_css': css, 'accent_color': color}


def _pluralize_unit(unit, qty):
    try:
        n = float(qty)
    except (TypeError, ValueError):
        return unit
    if not unit or n == 1:
        return unit
    irregulars = {'loaf': 'loaves', 'bunch': 'bunches', 'box': 'boxes'}
    low = unit.lower()
    if low in irregulars:
        return irregulars[low]
    if low.startswith('can '):
        return 'cans ' + unit[4:]
    if low in ('count', 'lb', 'lbs'):
        return unit
    return unit + 's'


app.jinja_env.filters['pluralize_unit'] = _pluralize_unit

XLSX               = Path(__file__).parent / 'Food in Storage.xlsx'
DATA_DIR           = Path(__file__).parent / 'data'
ENV_FILE           = Path(__file__).parent / '.env'
BOT_SCRIPT         = Path(__file__).parent / 'discord_bot.py'
BOT_PID_FILE       = Path(__file__).parent / 'bot.pid'
BOT_LOG_FILE       = Path(__file__).parent / 'bot.log'
BACKUP_DIR         = Path(__file__).parent / 'backups'
STATE_DIR          = Path(__file__).parent / 'state'
BACKUP_STATE_FILE  = STATE_DIR / 'backup_state.json'
SHOPPING_LIST_FILE = STATE_DIR / 'shopping_list.json'
EXCLUSIONS_FILE    = STATE_DIR / 'exclusions.json'
UNITS_FILE         = STATE_DIR / 'units.json'
EXCLUDE_SHEETS = {'Minimums'}
MEAL_TYPES     = ['Breakfast', 'Lunch', 'Dinner', 'Snacks', 'Desserts']

_bot_process = None



def write_env(env: dict):
    lines = [f'{k}={v}' for k, v in env.items()]
    ENV_FILE.write_text('\n'.join(lines) + '\n', encoding='utf-8')


def bot_running():
    return _bot_process is not None and _bot_process.poll() is None


def _kill_stale_bot():
    """Kill any bot process left over from a previous Flask session."""
    if not BOT_PID_FILE.exists():
        return
    try:
        pid = int(BOT_PID_FILE.read_text().strip())
        import os, signal
        os.kill(pid, signal.SIGTERM)
    except (ValueError, OSError, AttributeError):
        pass
    try:
        BOT_PID_FILE.unlink()
    except OSError:
        pass


def start_bot():
    global _bot_process
    if bot_running():
        _bot_process.terminate()
        _bot_process.wait()
    _kill_stale_bot()
    token = read_env().get('DISCORD_TOKEN', '')
    if not token:
        return False
    log = open(BOT_LOG_FILE, 'a', encoding='utf-8')
    log.write(f'\n--- Bot started {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} ---\n')
    log.flush()
    import os as _os
    bot_env = _os.environ.copy()
    bot_env['DISCORD_TOKEN'] = token
    _bot_process = subprocess.Popen(
        [sys.executable, str(BOT_SCRIPT)],
        cwd=str(BOT_SCRIPT.parent),
        stdout=log, stderr=log,
        env=bot_env,
    )
    BOT_PID_FILE.write_text(str(_bot_process.pid))
    return True


def stop_bot():
    global _bot_process
    if bot_running():
        _bot_process.terminate()
        _bot_process.wait()
    _bot_process = None
    _kill_stale_bot()


def init_wb():
    """Create a blank workbook with just the Minimums sheet if the file doesn't exist."""
    if not XLSX.exists():
        from openpyxl import Workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        ws = wb.create_sheet('Minimums')
        ws.cell(row=1, column=1).value = 'Item'
        ws.cell(row=1, column=2).value = 'Min Qty'
        wb.save(XLSX)


def get_location_sheets():
    init_wb()
    wb = load_workbook(XLSX)
    return [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]


def load_wb():
    init_wb()
    return load_workbook(XLSX)


def load_backup_state():
    try:
        return json.loads(BACKUP_STATE_FILE.read_text(encoding='utf-8')) if BACKUP_STATE_FILE.exists() else {}
    except Exception:
        return {}


def save_backup_state(state):
    BACKUP_STATE_FILE.write_text(json.dumps(state), encoding='utf-8')


def _do_backup(env):
    """Write a timestamped zip and trim to max. Does not touch state."""
    BACKUP_DIR.mkdir(exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = BACKUP_DIR / f'pantry_backup_{ts}.zip'
    with zipfile.ZipFile(dest, 'w', zipfile.ZIP_DEFLATED) as zf:
        if XLSX.exists():
            zf.write(XLSX, XLSX.name)
        if UNITS_FILE.exists():
            zf.write(UNITS_FILE, UNITS_FILE.name)
        if EXCLUSIONS_FILE.exists():
            zf.write(EXCLUSIONS_FILE, EXCLUSIONS_FILE.name)
        if DATA_DIR.exists():
            for jf in DATA_DIR.glob('*.json'):
                zf.write(jf, f'data/{jf.name}')
    max_backups = max(1, int(env.get('BACKUP_MAX', '10')))
    all_files = sorted(
        list(BACKUP_DIR.glob('*.zip')) + list(BACKUP_DIR.glob('*.xlsx')),
        key=lambda p: p.stat().st_mtime
    )
    for old in all_files[:-max_backups]:
        old.unlink(missing_ok=True)


def backup_wb(force=False):
    """Conditionally back up based on configured mode. force=True always writes."""
    if not XLSX.exists():
        return
    env   = read_env()
    mode  = env.get('BACKUP_MODE', 'actions')
    state = load_backup_state()
    state['action_count'] = state.get('action_count', 0) + 1

    do_backup = force
    if not do_backup:
        if mode == 'actions':
            n = max(1, int(env.get('BACKUP_EVERY_N', '10')))
            do_backup = state['action_count'] >= n
        elif mode == 'interval':
            hrs      = max(0.01, float(env.get('BACKUP_INTERVAL_HRS', '24')))
            last_ts  = state.get('last_backup_ts', 0)
            do_backup = (datetime.now().timestamp() - last_ts) >= (hrs * 3600)

    if do_backup:
        _do_backup(env)
        state['action_count']  = 0
        state['last_backup_ts'] = datetime.now().timestamp()

    save_backup_state(state)


def get_backups():
    if not BACKUP_DIR.exists():
        return []
    files = list(BACKUP_DIR.glob('*.zip')) + list(BACKUP_DIR.glob('*.xlsx'))
    result = []
    for p in sorted(files, key=lambda p: p.stat().st_mtime, reverse=True):
        stat = p.stat()
        kb   = stat.st_size / 1024
        result.append({
            'name':     p.name,
            'size':     f'{kb:.0f} KB' if kb >= 1 else f'{stat.st_size} B',
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%b %d %Y, %I:%M %p'),
        })
    return result


def get_backup_info():
    """Return (last_label, action_count) for the settings UI."""
    state = load_backup_state()
    count = state.get('action_count', 0)
    last  = state.get('last_backup_ts', 0)
    if last:
        ago_s = datetime.now().timestamp() - last
        ago_h = ago_s / 3600
        if ago_h < 1:
            last_str = f'{int(ago_s / 60)} min ago'
        elif ago_h < 24:
            last_str = f'{ago_h:.1f} hrs ago'
        else:
            last_str = f'{ago_h / 24:.1f} days ago'
        last_label = f'Last backup: {last_str}'
    else:
        last_label = 'No backup taken yet'
    return last_label, count


def save_wb(wb):
    backup_wb()
    dedup_workbook(wb)
    wb.save(XLSX)


def dedup_workbook(wb):
    """Merge rows with the same item name AND unit within each sheet. Returns True if changes were made."""
    changed = False
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        seen = {}  # (item_lower, unit_lower) -> first row_idx
        to_clear = []
        for r in range(2, ws.max_row + 1):
            item = ws.cell(row=r, column=1).value
            if not item or str(item).startswith('='):
                continue
            unit = str(ws.cell(row=r, column=3).value or '').strip().lower()
            key = (str(item).strip().lower(), unit)
            if key in seen:
                first = seen[key]
                try:
                    q1 = float(str(ws.cell(row=first, column=2).value or 0))
                    q2 = float(str(ws.cell(row=r,     column=2).value or 0))
                    merged = q1 + q2
                    ws.cell(row=first, column=2).value = int(merged) if merged == int(merged) else merged
                except (ValueError, TypeError):
                    pass
                to_clear.append(r)
                changed = True
            else:
                seen[key] = r
        for r in reversed(to_clear):
            ws.delete_rows(r)
    return changed


def get_all_rows(wb):
    rows = []
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            item = vals[0]
            if any(v for v in vals) and item and not str(item).startswith('='):
                rows.append((sheet_name, r, dict(zip(COLS, vals))))
    return rows


def get_location_info():
    wb = load_wb()
    result = []
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value)
        result.append((sheet_name, count))
    return result


def get_minimums():
    wb = load_wb()
    if 'Minimums' not in wb.sheetnames:
        return []
    ws = wb['Minimums']
    result = []
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=1).value
        qty  = ws.cell(row=r, column=2).value
        if item:
            result.append((str(item), qty, r))
    return sorted(result, key=lambda x: x[0].lower())


def set_minimum(item_title, qty):
    wb = load_wb()
    if 'Minimums' not in wb.sheetnames:
        ws = wb.create_sheet('Minimums')
        ws.cell(row=1, column=1).value = 'Item'
        ws.cell(row=1, column=2).value = 'Min Qty'
    else:
        ws = wb['Minimums']
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or '').lower() == item_title.lower():
            ws.cell(row=r, column=2).value = qty
            save_wb(wb)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = item_title
    ws.cell(row=row, column=2).value = qty
    save_wb(wb)


def delete_minimum(item_title):
    wb = load_wb()
    if 'Minimums' not in wb.sheetnames:
        return
    ws = wb['Minimums']
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or '').lower() == item_title.lower():
            ws.delete_rows(r)
            save_wb(wb)
            return


def load_shopping_list():
    if not SHOPPING_LIST_FILE.exists():
        return []
    try:
        return json.loads(SHOPPING_LIST_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def save_shopping_list(items):
    SHOPPING_LIST_FILE.write_text(json.dumps(items, indent=2), encoding='utf-8')


def load_exclusions():
    if not EXCLUSIONS_FILE.exists():
        return []
    try:
        return json.loads(EXCLUSIONS_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def save_exclusions(items):
    EXCLUSIONS_FILE.write_text(json.dumps(items, indent=2), encoding='utf-8')


def load_units():
    if not UNITS_FILE.exists():
        return []
    try:
        return json.loads(UNITS_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def save_units(units):
    UNITS_FILE.write_text(json.dumps(sorted(units, key=str.lower), indent=2), encoding='utf-8')


def get_unit_info():
    units = load_units()
    if not XLSX.exists():
        return [(u, 0) for u in units]
    wb = load_wb()
    counts = {u.lower(): 0 for u in units}
    for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=3).value
            if cell_val:
                key = str(cell_val).strip().lower()
                if key in counts:
                    counts[key] += 1
    return [(u, counts[u.lower()]) for u in units]


def is_excluded(item_name):
    name = item_name.strip().lower()
    for ex in load_exclusions():
        ex_name = ex.get('name', '').lower()
        if name == ex_name or name in ex_name or ex_name in name:
            return True
    return False


BASE_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Pantri</title>
<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>🥘</text></svg>">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
<style>
  {{ theme_css }}
  body { background: #f9f6fc; font-family: 'Segoe UI', sans-serif; }
  .navbar { background: var(--purple) !important; }
  .navbar-brand, .navbar .nav-link { color: white !important; font-weight: 600; }
  .navbar .nav-link:hover { color: var(--light-purple) !important; }
  .nav-tabs .nav-link { color: var(--dark-purple) !important; font-weight: 600; background: transparent; border-color: var(--light-purple) var(--light-purple) transparent; }
  .nav-tabs .nav-link:hover:not(.active) { background: var(--light-purple); color: var(--dark-purple) !important; }
  .nav-tabs .nav-link.active { color: white !important; background: var(--purple); border-color: var(--purple) var(--purple) var(--purple); }
  .nav-tabs { border-bottom: 2px solid var(--light-purple); }
  .card { border: 1.5px solid var(--light-purple); border-radius: 12px; }
  .card-header { background: var(--purple); color: white; border-radius: 10px 10px 0 0 !important; font-weight: 700; }
  .btn-primary { background: var(--purple); border-color: var(--dark-purple); }
  .btn-primary:hover { background: var(--dark-purple); }
  .btn-outline-primary { color: var(--purple); border-color: var(--purple); }
  .btn-outline-primary:hover { background: var(--purple); color: white; }
  table thead { background: var(--dark-purple); color: white; }
  table tbody tr:hover { background: var(--light-purple); }
  .search-bar { border: 2px solid var(--light-purple); border-radius: 8px; }
  .search-bar:focus { border-color: var(--purple); box-shadow: 0 0 0 0.2rem rgba(107,45,139,.15); }
  h1 small { font-size: 0.5em; color: var(--light-purple); }
  .ingredient-row { display: flex; gap: 8px; margin-bottom: 6px; align-items: center; }
  .ingredient-row input { flex: 1; }
  .ingredient-row .amount-input { flex: 0 0 120px; }
  .instruction-row { display: flex; gap: 8px; margin-bottom: 6px; align-items: flex-start; }
  .instruction-row span { padding-top: 8px; font-weight: bold; color: var(--purple); min-width: 24px; }
  .instruction-row textarea { flex: 1; }
</style>
</head>
<body>
<nav class="navbar navbar-expand-lg mb-4">
  <div class="container">
    <a class="navbar-brand" href="/">Pantri</a>
    <div class="navbar-nav ms-auto">
      <a class="nav-link" href="/">Inventory</a>
      <a class="nav-link" href="/recipes">Recipes</a>
      <a class="nav-link" href="/shopping">Shopping</a>
      <a class="nav-link" href="/settings">⚙ Settings</a>
    </div>
  </div>
</nav>
<div class="container pb-5">
  {% block content %}{% endblock %}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

INDEX_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<h1 class="mb-3">Inventory</h1>

<ul class="nav nav-tabs mb-4">
  <li class="nav-item">
    <a class="nav-link {% if tab != 'add' %}active{% endif %}" href="/">
      Items <span class="badge ms-1" style="background:var(--light-purple);color:var(--dark-purple);">{{ total }}</span>
    </a>
  </li>
  <li class="nav-item">
    <a class="nav-link {% if tab == 'add' %}active{% endif %}" href="/?tab=add">Add Items</a>
  </li>
</ul>

{% if tab != 'add' %}

  {% if not locations %}
  <div class="card">
    <div class="card-body text-center py-5">
      <div style="font-size:3rem;margin-bottom:1rem;">📦</div>
      <h4 style="color:var(--dark-purple);">No storage locations set up yet</h4>
      <p class="text-muted mb-4">Add your first location (e.g. Pantry, Fridge, Garage Shelf) before tracking inventory.</p>
      <a href="/settings?tab=locations" class="btn btn-primary px-4">Go to Settings → Locations</a>
    </div>
  </div>
  {% else %}

  <div class="card mb-4">
    <div class="card-body">
      <form method="get" class="row g-2">
        <div class="col-md-5">
          <input name="q" class="form-control search-bar" placeholder="Search items..." value="{{ q }}">
        </div>
        <div class="col-md-4">
          <select name="loc" class="form-select search-bar">
            <option value="">All Locations</option>
            {% for l in locations %}
            <option value="{{ l }}" {% if loc == l %}selected{% endif %}>{{ l }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <button type="submit" class="btn btn-primary w-100">Filter</button>
        </div>
      </form>
    </div>
  </div>
  <div class="card">
    <div class="card-header">Items</div>
    <div class="table-responsive">
      <table class="table table-hover mb-0">
        <thead>
          <tr><th>Item</th><th>Qty</th><th>Unit</th><th>Location</th><th></th></tr>
        </thead>
        <tbody>
          {% for i, row in items %}
          <tr>
            <td><strong>{{ row.Item }}</strong></td>
            <td>{{ row.Quantity or '' }}</td>
            <td>{{ row.Unit | pluralize_unit(row.Quantity) }}</td>
            <td>{{ row.Location or '' }}</td>
            <td>
              <a href="/edit/{{ i }}" class="btn btn-sm btn-outline-primary me-1">Edit</a>
              <a href="/delete/{{ i }}" class="btn btn-sm btn-outline-danger"
                 onclick="return confirm('Remove this item?')">Del</a>
            </td>
          </tr>
          {% endfor %}
          {% if not items %}
          <tr><td colspan="5" class="text-center text-muted py-4">No items found.</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

  {% endif %}

{% else %}

  {% if not locations %}
  <div class="card">
    <div class="card-body text-center py-5">
      <div style="font-size:3rem;margin-bottom:1rem;">📦</div>
      <h4 style="color:var(--dark-purple);">Add a location first</h4>
      <p class="text-muted mb-4">You need at least one storage location before you can add items.</p>
      <a href="/settings?tab=locations" class="btn btn-primary px-4">Go to Settings → Locations</a>
    </div>
  </div>
  {% else %}

  <div class="row g-4 align-items-start">
    <div class="col-md-6">
      <div class="card h-100">
        <div class="card-header">Add Single Item</div>
        <div class="card-body">
          <form method="post" action="/add">
            <div class="mb-3">
              <label class="form-label fw-bold">Item</label>
              <input name="Item" class="form-control" placeholder="e.g. Black Beans" required>
            </div>
            <div class="mb-3">
              <label class="form-label fw-bold">Quantity</label>
              <input name="Quantity" class="form-control" placeholder="e.g. 3">
            </div>
            <div class="mb-3">
              <label class="form-label fw-bold">Unit</label>
              <select name="Unit" class="form-select">
                <option value="">— none —</option>
                {% for u in units %}
                <option>{{ u }}</option>
                {% endfor %}
              </select>
            </div>
            <div class="mb-3">
              <label class="form-label fw-bold">Location</label>
              <select name="Location" class="form-select">
                {% for l in locations %}
                <option>{{ l }}</option>
                {% endfor %}
              </select>
            </div>
            <button type="submit" class="btn btn-primary w-100">Add Item</button>
          </form>
        </div>
      </div>
    </div>
    <div class="col-md-6">
      <div class="card h-100">
        <div class="card-header">Bulk Add</div>
        <div class="card-body d-flex flex-column">
          <p class="small text-muted mb-2">
            One item per line: <code>Item, Qty, Unit, Location</code><br>
            Qty / Unit / Location are optional. Examples:<br>
            <code>Black Beans, 3, 15oz can, Pantry</code><br>
            <code>Olive Oil, 1, bottle</code><br>
            <code>Salt</code>
          </p>
          <form method="post" action="/add/bulk" class="d-flex flex-column flex-grow-1">
            <div class="mb-3 flex-grow-1">
              <textarea name="bulk" class="form-control font-monospace h-100" rows="10"
                        style="min-height:200px"
                        placeholder="Black Beans, 3, 15oz can, Pantry&#10;Olive Oil, 1, bottle&#10;Salt"></textarea>
            </div>
            <div class="mb-3">
              <label class="form-label fw-bold">Default Location</label>
              <select name="default_location" class="form-select">
                {% for l in locations %}
                <option>{{ l }}</option>
                {% endfor %}
              </select>
            </div>
            <button type="submit" class="btn btn-primary w-100">Add All</button>
          </form>
          {% if bulk_results %}
          <hr>
          <h6>Results:</h6>
          <ul class="small mb-0">
            {% for msg in bulk_results %}
            <li>{{ msg }}</li>
            {% endfor %}
          </ul>
          {% endif %}
        </div>
      </div>
    </div>
  </div>

  {% endif %}

{% endif %}
""")

EDIT_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<h1>{{ title }}</h1>
<div class="card mt-3" style="max-width:500px">
  <div class="card-header">{{ title }}</div>
  <div class="card-body">
    <form method="post">
      <div class="mb-3">
        <label class="form-label fw-bold">Item</label>
        <input name="Item" class="form-control" value="{{ values.get('Item', '') }}" required>
      </div>
      <div class="mb-3">
        <label class="form-label fw-bold">Quantity</label>
        <input name="Quantity" class="form-control" value="{{ values.get('Quantity', '') }}">
      </div>
      <div class="mb-3">
        <label class="form-label fw-bold">Unit</label>
        <select name="Unit" class="form-select">
          <option value="">— none —</option>
          {% for u in units %}
          <option {% if values.get('Unit', '') == u %}selected{% endif %}>{{ u }}</option>
          {% endfor %}
          {% if values.get('Unit') and values.get('Unit') not in units %}
          <option selected>{{ values.get('Unit') }}</option>
          {% endif %}
        </select>
      </div>
      <div class="mb-3">
        <label class="form-label fw-bold">Location</label>
        <select name="Location" class="form-select">
          {% for l in locations %}
          <option value="{{ l }}" {% if values.get('Location') == l %}selected{% endif %}>{{ l }}</option>
          {% endfor %}
        </select>
      </div>
      <div class="d-flex gap-2">
        <button type="submit" class="btn btn-primary">Save</button>
        <a href="/" class="btn btn-outline-secondary">Cancel</a>
      </div>
    </form>
  </div>
</div>
""")



RECIPES_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<div class="d-flex justify-content-between align-items-center mb-3">
  <h1>Recipes</h1>
  <div class="d-flex gap-2">
    <a href="/recipes/download-pdf" class="btn btn-outline-primary">⬇ Download PDF</a>
    <a href="/recipe/add" class="btn btn-primary">+ Add Recipe</a>
  </div>
</div>

<ul class="nav nav-tabs mb-4">
  <li class="nav-item">
    <a class="nav-link {% if tab != 'canmake' %}active{% endif %}" href="/recipes">
      Recipe Library <span class="badge ms-1" style="background:var(--light-purple);color:var(--dark-purple);">{{ recipes|length }}</span>
    </a>
  </li>
  <li class="nav-item">
    <a class="nav-link {% if tab == 'canmake' %}active{% endif %}" href="/recipes?tab=canmake">Can I Make?</a>
  </li>
</ul>

{% if tab != 'canmake' %}

  <div class="card mb-4">
    <div class="card-body">
      <form method="get" class="row g-2">
        <div class="col-md-5">
          <input name="q" class="form-control search-bar" placeholder="Search recipes..." value="{{ q }}">
        </div>
        <div class="col-md-4">
          <select name="meal" class="form-select search-bar">
            <option value="">All Meal Types</option>
            {% for m in meal_types %}
            <option value="{{ m }}" {% if meal == m %}selected{% endif %}>{{ m }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <button type="submit" class="btn btn-primary w-100">Filter</button>
        </div>
      </form>
    </div>
  </div>
  <div class="card">
    <div class="card-header">Recipe Library</div>
    <div class="table-responsive">
      <table class="table table-hover mb-0">
        <thead><tr><th>Name</th><th>Meal Type</th><th>Prep</th><th>Serves</th><th></th></tr></thead>
        <tbody>
          {% for r in recipes %}
          <tr>
            <td><a href="/recipe/{{ r.slug }}" class="fw-bold text-decoration-none" style="color:#4A1E6B;">{{ r.name }}</a></td>
            <td>{{ r.meal_type }}</td>
            <td>{{ r.prep_time }}</td>
            <td>{{ r.servings }}</td>
            <td>
              <a href="/recipe/delete/{{ r.slug }}" class="btn btn-sm btn-outline-danger"
                 onclick="return confirm('Delete {{ r.name }}?')">Del</a>
            </td>
          </tr>
          {% endfor %}
          {% if not recipes %}
          <tr><td colspan="5" class="text-center text-muted py-4">No recipes found.</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

{% else %}

  <div class="card mb-4">
    <div class="card-body">
      <form method="get" action="/recipes" class="row g-2 align-items-end">
        <input type="hidden" name="tab" value="canmake">
        <div class="col-12 mb-1">
          <span class="fw-bold small me-2">Meal Type:</span>
          <a href="/recipes?tab=canmake&ingredient={{ ingredient }}"
             class="btn btn-sm me-1 {% if not cm_meal %}btn-primary{% else %}btn-outline-primary{% endif %}">All</a>
          {% for m in meal_types %}
          <a href="/recipes?tab=canmake&cm_meal={{ m }}&ingredient={{ ingredient }}"
             class="btn btn-sm me-1 {% if cm_meal == m %}btn-primary{% else %}btn-outline-primary{% endif %}">{{ m }}</a>
          {% endfor %}
        </div>
        <div class="col-md-6">
          <label class="form-label fw-bold small mb-1">Contains Ingredient</label>
          <div class="d-flex gap-2">
            <input name="ingredient" class="form-control search-bar"
                   placeholder="e.g. tuna, chicken, oats…" value="{{ ingredient }}">
            {% if cm_meal %}<input type="hidden" name="cm_meal" value="{{ cm_meal }}">{% endif %}
            <button type="submit" class="btn btn-primary px-3">Filter</button>
            {% if cm_meal or ingredient %}
            <a href="/recipes?tab=canmake" class="btn btn-outline-secondary px-3">Clear</a>
            {% endif %}
          </div>
        </div>
      </form>
    </div>
  </div>

  <div class="row row-cols-1 row-cols-md-2 g-4">
    {% for recipe in cm_recipes %}
    <div class="col">
      <div class="card h-100">
        <div class="card-header">{{ recipe.name }}</div>
        <div class="card-body">
          <div class="mb-2">
            <span class="{{ recipe.status_class }}">{{ recipe.status_label }}</span>
            <span class="text-muted ms-2 small">{{ recipe.meal_type }} · {{ recipe.prep_time }}</span>
          </div>
          {% if recipe.missing %}
          <p class="small text-muted mb-1">Missing:</p>
          <ul class="small mb-0">
            {% for m in recipe.missing %}<li>{{ m }}</li>{% endfor %}
          </ul>
          {% endif %}
        </div>
      </div>
    </div>
    {% endfor %}
    {% if not cm_recipes %}
    <div class="col-12"><p class="text-muted">No recipes match those filters.</p></div>
    {% endif %}
  </div>

{% endif %}
""")

ADD_RECIPE_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<h1>{{ page_title }}</h1>

{% if not edit_mode %}
<div class="card mb-3" style="border-color:var(--purple);">
  <div class="card-body">
    <label class="form-label fw-bold mb-1">Import from URL</label>
    <p class="text-muted small mb-2">Paste any recipe URL (AllRecipes, Food Network, most food blogs, etc.) and we'll fill in the form automatically.</p>
    <div class="d-flex gap-2">
      <input id="import-url" type="url" class="form-control search-bar" placeholder="https://www.allrecipes.com/recipe/...">
      <button type="button" id="import-btn" class="btn btn-primary px-4" onclick="importFromUrl()">Import</button>
    </div>
    <div id="import-error" class="text-danger small mt-2 d-none"></div>
    <div id="import-success" class="text-success small mt-2 d-none">✅ Recipe imported — review the fields below and save.</div>
  </div>
</div>
{% endif %}

<div class="card mt-3">
  <div class="card-header">{{ page_title }}</div>
  <div class="card-body">
    <form method="post" id="recipeForm">

      <div class="row g-3 mb-3">
        <div class="col-md-6">
          <label class="form-label fw-bold">Recipe Name</label>
          <input name="name" class="form-control" required value="{{ rv.name }}">
        </div>
        <div class="col-md-3">
          <label class="form-label fw-bold">Meal Type</label>
          <select name="meal_type" class="form-select">
            {% for m in meal_types %}
            <option {% if rv.meal_type == m %}selected{% endif %}>{{ m }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <label class="form-label fw-bold">Servings</label>
          <input name="servings" type="number" min="1" class="form-control" value="{{ rv.servings or 4 }}">
        </div>
      </div>

      <div class="row g-3 mb-3">
        <div class="col-md-4">
          <label class="form-label fw-bold">Prep Time</label>
          <input name="prep_time" class="form-control" placeholder="e.g. 15 minutes" value="{{ rv.prep_time }}">
        </div>
        <div class="col-md-4">
          <label class="form-label fw-bold">Cook Time</label>
          <input name="cook_time" class="form-control" placeholder="e.g. 30 minutes" value="{{ rv.cook_time }}">
        </div>
      </div>

      <div class="card mb-3">
        <div class="card-header" style="background:#f3eaf8;color:#4A1E6B;">Macros per Serving (optional)</div>
        <div class="card-body">
          <div class="row g-3">
            <div class="col-6 col-md-3">
              <label class="form-label">Calories</label>
              <input name="calories" type="number" class="form-control" placeholder="0" value="{{ rv.calories }}">
            </div>
            <div class="col-6 col-md-3">
              <label class="form-label">Protein (g)</label>
              <input name="protein" type="number" class="form-control" placeholder="0" value="{{ rv.protein }}">
            </div>
            <div class="col-6 col-md-3">
              <label class="form-label">Carbs (g)</label>
              <input name="carbs" type="number" class="form-control" placeholder="0" value="{{ rv.carbs }}">
            </div>
            <div class="col-6 col-md-3">
              <label class="form-label">Fat (g)</label>
              <input name="fat" type="number" class="form-control" placeholder="0" value="{{ rv.fat }}">
            </div>
          </div>
        </div>
      </div>

      <div class="mb-3">
        <label class="form-label fw-bold">Ingredients</label>
        <div id="ingredients">
          {% for ing in rv.ingredients %}
          <div class="ingredient-row">
            <input class="form-control amount-input" name="amount[]" placeholder="Amount (e.g. 1.5 lbs)" value="{{ ing.amount }}">
            <input class="form-control" name="ingredient[]" placeholder="Ingredient name" value="{{ ing.item }}">
            <button type="button" class="btn btn-outline-danger btn-sm" onclick="removeRow(this)">✕</button>
          </div>
          {% else %}
          <div class="ingredient-row">
            <input class="form-control amount-input" name="amount[]" placeholder="Amount (e.g. 1.5 lbs)">
            <input class="form-control" name="ingredient[]" placeholder="Ingredient name">
            <button type="button" class="btn btn-outline-danger btn-sm" onclick="removeRow(this)">✕</button>
          </div>
          {% endfor %}
        </div>
        <button type="button" class="btn btn-outline-primary btn-sm mt-2" onclick="addIngredient()">+ Ingredient</button>
      </div>

      <div class="mb-3">
        <label class="form-label fw-bold">Instructions</label>
        <div id="instructions">
          {% for step in rv.instructions %}
          <div class="instruction-row">
            <span>{{ loop.index }}.</span>
            <textarea class="form-control" name="instruction[]" rows="2">{{ step }}</textarea>
            <button type="button" class="btn btn-outline-danger btn-sm mt-1" onclick="removeRow(this)">✕</button>
          </div>
          {% else %}
          <div class="instruction-row">
            <span>1.</span>
            <textarea class="form-control" name="instruction[]" rows="2" placeholder="Step 1..."></textarea>
            <button type="button" class="btn btn-outline-danger btn-sm mt-1" onclick="removeRow(this)">✕</button>
          </div>
          {% endfor %}
        </div>
        <button type="button" class="btn btn-outline-primary btn-sm mt-2" onclick="addInstruction()">+ Step</button>
      </div>

      <div class="card mb-3">
        <div class="card-header" style="background:#f3eaf8;color:#4A1E6B;">Substitutions (optional)</div>
        <div class="card-body">
          <div class="row g-3">
            <div class="col-md-6">
              <label class="form-label">Dairy-Free</label>
              <input name="sub_dairy_free" class="form-control" placeholder="e.g. Use oat milk" value="{{ rv.sub_dairy_free }}">
            </div>
            <div class="col-md-6">
              <label class="form-label">Gluten-Free</label>
              <input name="sub_gluten_free" class="form-control" placeholder="e.g. Use GF flour" value="{{ rv.sub_gluten_free }}">
            </div>
            <div class="col-md-6">
              <label class="form-label">Vegan</label>
              <input name="sub_vegan" class="form-control" placeholder="e.g. Use flax eggs" value="{{ rv.sub_vegan }}">
            </div>
            <div class="col-md-6">
              <label class="form-label">Vegetarian</label>
              <input name="sub_vegetarian" class="form-control" placeholder="e.g. Skip bacon" value="{{ rv.sub_vegetarian }}">
            </div>
            <div class="col-md-6">
              <label class="form-label">Low-Carb</label>
              <input name="sub_low_carb" class="form-control" placeholder="e.g. Use cauliflower" value="{{ rv.sub_low_carb }}">
            </div>
          </div>
        </div>
      </div>

      <div class="mb-3">
        <label class="form-label fw-bold">Notes (optional)</label>
        <textarea name="notes" class="form-control" rows="2">{{ rv.notes }}</textarea>
      </div>

      <div class="d-flex gap-2">
        <button type="submit" class="btn btn-primary px-4">Save Recipe</button>
        <a href="/recipes" class="btn btn-outline-secondary">Cancel</a>
      </div>
    </form>
  </div>
</div>

<script>
function addIngredient(amount='', item='') {
  const container = document.getElementById('ingredients');
  const row = document.createElement('div');
  row.className = 'ingredient-row';
  row.innerHTML = `
    <input class="form-control amount-input" name="amount[]" placeholder="Amount (e.g. 1.5 lbs)" value="${amount.replace(/"/g,'&quot;')}">
    <input class="form-control" name="ingredient[]" placeholder="Ingredient name" value="${item.replace(/"/g,'&quot;')}">
    <button type="button" class="btn btn-outline-danger btn-sm" onclick="removeRow(this)">✕</button>`;
  container.appendChild(row);
}
function addInstruction(text='') {
  const container = document.getElementById('instructions');
  const num = container.children.length + 1;
  const row = document.createElement('div');
  row.className = 'instruction-row';
  row.innerHTML = `
    <span>${num}.</span>
    <textarea class="form-control" name="instruction[]" rows="2" placeholder="Step ${num}...">${text.replace(/</g,'&lt;')}</textarea>
    <button type="button" class="btn btn-outline-danger btn-sm mt-1" onclick="removeRow(this)">✕</button>`;
  container.appendChild(row);
}
function removeRow(btn) {
  btn.closest('.ingredient-row, .instruction-row').remove();
  document.querySelectorAll('#instructions .instruction-row').forEach((row, i) => {
    row.querySelector('span').textContent = (i + 1) + '.';
  });
}
async function importFromUrl() {
  const url = document.getElementById('import-url').value.trim();
  const errEl = document.getElementById('import-error');
  const okEl  = document.getElementById('import-success');
  const btn   = document.getElementById('import-btn');
  errEl.classList.add('d-none');
  okEl.classList.add('d-none');
  if (!url) return;
  btn.disabled = true;
  btn.textContent = 'Importing…';
  try {
    const resp = await fetch('/recipe/import?url=' + encodeURIComponent(url));
    const data = await resp.json();
    if (data.error) { errEl.textContent = data.error; errEl.classList.remove('d-none'); return; }

    const f = document.getElementById('recipeForm');
    f.querySelector('[name=name]').value     = data.name     || '';
    f.querySelector('[name=servings]').value = data.servings || 4;
    f.querySelector('[name=prep_time]').value = data.prep_time || '';
    f.querySelector('[name=cook_time]').value = data.cook_time || '';
    f.querySelector('[name=calories]').value  = data.calories  || '';
    f.querySelector('[name=protein]').value   = data.protein   || '';
    f.querySelector('[name=carbs]').value     = data.carbs     || '';
    f.querySelector('[name=fat]').value       = data.fat       || '';
    f.querySelector('[name=notes]').value     = data.notes     || '';

    const mtSel = f.querySelector('[name=meal_type]');
    [...mtSel.options].forEach(o => { if (o.value === data.meal_type) mtSel.value = data.meal_type; });

    document.getElementById('ingredients').innerHTML = '';
    (data.ingredients || []).forEach(i => addIngredient(i.amount, i.item));
    if (!(data.ingredients || []).length) addIngredient();

    document.getElementById('instructions').innerHTML = '';
    (data.instructions || []).forEach(s => addInstruction(s));
    if (!(data.instructions || []).length) addInstruction();

    okEl.classList.remove('d-none');
    window.scrollTo(0, 0);
  } catch(e) {
    errEl.textContent = 'Something went wrong — check the URL and try again.';
    errEl.classList.remove('d-none');
  } finally {
    btn.disabled = false;
    btn.textContent = 'Import';
  }
}
</script>
""")

SETTINGS_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<h1 class="mb-3">Settings</h1>

<ul class="nav nav-tabs mb-4" id="settingsTabs" role="tablist">
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'locations' or active_tab not in ('units', 'minimums', 'exclusions', 'backups', 'appearance', 'discord') %}active{% endif %}"
            id="locations-tab" data-bs-toggle="tab" data-bs-target="#locations"
            type="button" role="tab">Locations</button>
  </li>
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'units' %}active{% endif %}"
            id="units-tab" data-bs-toggle="tab" data-bs-target="#units"
            type="button" role="tab">Units</button>
  </li>
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'minimums' %}active{% endif %}"
            id="minimums-tab" data-bs-toggle="tab" data-bs-target="#minimums"
            type="button" role="tab">Minimums</button>
  </li>
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'exclusions' %}active{% endif %}"
            id="exclusions-tab" data-bs-toggle="tab" data-bs-target="#exclusions"
            type="button" role="tab">Exclusions</button>
  </li>
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'backups' %}active{% endif %}"
            id="backups-tab" data-bs-toggle="tab" data-bs-target="#backups"
            type="button" role="tab">Backups</button>
  </li>
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'appearance' %}active{% endif %}"
            id="appearance-tab" data-bs-toggle="tab" data-bs-target="#appearance"
            type="button" role="tab">Appearance</button>
  </li>
  <li class="nav-item" role="presentation">
    <button class="nav-link {% if active_tab == 'discord' %}active{% endif %}"
            id="discord-tab" data-bs-toggle="tab" data-bs-target="#discord"
            type="button" role="tab">Discord Bot</button>
  </li>
</ul>

<div class="tab-content">

  <!-- ── Locations Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'locations' or active_tab not in ('units', 'minimums', 'exclusions', 'backups', 'appearance', 'discord') %}show active{% endif %}"
       id="locations" role="tabpanel">
    <div class="row g-4">
      <div class="col-md-8">
        <div class="card">
          <div class="card-header">Storage Locations <small class="fw-normal ms-2" style="font-size:.8em;opacity:.7">drag to reorder</small></div>
          <div class="table-responsive">
            <table class="table table-hover mb-0">
              <thead><tr><th style="width:2rem"></th><th>Name</th><th>Items</th><th></th></tr></thead>
              <tbody id="locations-sortable">
                {% for name, count in locations %}
                <tr data-name="{{ name }}">
                  <td class="text-muted" style="cursor:grab;user-select:none;font-size:1.1rem">&#9776;</td>
                  <td><strong>{{ name }}</strong></td>
                  <td>{{ count }}</td>
                  <td>
                    {% if count == 0 %}
                    <a href="/locations/delete/{{ name }}" class="btn btn-sm btn-outline-danger"
                       onclick="return confirm('Delete location {{ name }}?')">Delete</a>
                    {% else %}
                    <span class="text-muted small">{{ count }} item{{ 's' if count != 1 else '' }} — clear first to delete</span>
                    {% endif %}
                  </td>
                </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>
        </div>
      </div>
      <div class="col-md-4">
        <div class="card">
          <div class="card-header">Add Location</div>
          <div class="card-body">
            <form method="post" action="/locations/add">
              <div class="mb-3">
                <label class="form-label fw-bold">Name</label>
                <input name="name" class="form-control" placeholder="e.g. Garage Shelf" required>
              </div>
              <button type="submit" class="btn btn-primary w-100">Add</button>
            </form>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Units Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'units' %}show active{% endif %}"
       id="units" role="tabpanel">
    <div class="row g-4">
      <div class="col-md-8">
        <div class="card">
          <div class="card-header">Unit List <small class="fw-normal ms-2" style="font-size:.8em;opacity:.8">used in add / edit dropdowns &nbsp;·&nbsp; drag to reorder</small></div>
          <div class="table-responsive">
            <table class="table table-hover mb-0">
              <thead><tr><th style="width:2rem"></th><th>Unit</th><th>Items using it</th><th></th></tr></thead>
              <tbody id="units-sortable">
                {% for unit, count in units %}
                <tr data-unit="{{ unit }}">
                  <td class="text-muted" style="cursor:grab;user-select:none;font-size:1.1rem">&#9776;</td>
                  <td><strong>{{ unit }}</strong></td>
                  <td>{{ count }}</td>
                  <td class="text-end">
                    {% if count == 0 %}
                    <a href="/units/delete/{{ unit | urlencode }}" class="btn btn-sm btn-outline-danger"
                       onclick="return confirm('Delete unit {{ unit }}?')">Delete</a>
                    {% else %}
                    <span class="text-muted small">{{ count }} item{{ 's' if count != 1 else '' }} — in use</span>
                    {% endif %}
                  </td>
                </tr>
                {% endfor %}
                {% if not units %}
                <tr><td colspan="4" class="text-center text-muted py-4">No units defined yet.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        </div>
      </div>
      <div class="col-md-4">
        <div class="card">
          <div class="card-header">Add Unit</div>
          <div class="card-body">
            <p class="small text-muted mb-3">Add standardized unit names (e.g. can, bag, lb). These appear in the add/edit dropdowns so unit spelling stays consistent.</p>
            <form method="post" action="/units/add">
              <div class="mb-3">
                <label class="form-label fw-bold">Unit Name</label>
                <input name="unit" class="form-control" placeholder="e.g. can, bag, lb" required>
              </div>
              <button type="submit" class="btn btn-primary w-100">Add</button>
            </form>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Minimums Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'minimums' %}show active{% endif %}"
       id="minimums" role="tabpanel">
    <div class="row g-4">
      <div class="col-md-8">
        <div class="card">
          <div class="card-header">Current Minimums <small class="fw-normal ms-2" style="font-size:.8em;opacity:.8">click a number to edit</small></div>
          <div class="table-responsive">
            <table class="table table-hover mb-0">
              <thead><tr><th>Item</th><th>Min Qty</th><th></th></tr></thead>
              <tbody>
                {% for item, qty, _ in minimums %}
                <tr>
                  <td><strong>{{ item }}</strong></td>
                  <td>
                    <span class="qty-display" data-item="{{ item }}" title="Click to edit"
                          style="cursor:pointer;border-bottom:1px dashed var(--purple);padding-bottom:1px;">{{ qty }}</span>
                    <form class="qty-form d-none" method="post" action="/minimums/set" style="display:inline;">
                      <input type="hidden" name="item" value="{{ item }}">
                      <input type="number" name="qty" step="0.1" min="0" value="{{ qty }}"
                             class="form-control form-control-sm d-inline-block" style="width:80px;">
                      <button type="submit" class="btn btn-sm btn-primary ms-1">✓</button>
                      <button type="button" class="btn btn-sm btn-outline-secondary ms-1 qty-cancel">✕</button>
                    </form>
                  </td>
                  <td>
                    <a href="/minimums/delete/{{ item }}" class="btn btn-sm btn-outline-danger"
                       onclick="return confirm('Delete minimum for {{ item }}?')">Del</a>
                  </td>
                </tr>
                {% endfor %}
                {% if not minimums %}
                <tr><td colspan="3" class="text-center text-muted py-4">No minimums set.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        </div>
      </div>
      <div class="col-md-4">
        <div class="card">
          <div class="card-header">Add New Minimum</div>
          <div class="card-body">
            <form method="post" action="/minimums/set">
              <div class="mb-3">
                <label class="form-label fw-bold">Item</label>
                <input name="item" class="form-control" placeholder="e.g. Black Beans" required>
              </div>
              <div class="mb-3">
                <label class="form-label fw-bold">Min Qty</label>
                <input name="qty" type="number" step="0.1" min="0" class="form-control" required>
              </div>
              <button type="submit" class="btn btn-primary w-100">Save</button>
            </form>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Exclusions Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'exclusions' %}show active{% endif %}"
       id="exclusions" role="tabpanel">
    <div class="row g-4">
      <div class="col-md-8">
        <div class="card">
          <div class="card-header">Excluded Items <small class="fw-normal ms-2" style="font-size:.8em;opacity:.8">never auto-decremented from pantry</small></div>
          <div class="table-responsive">
            <table class="table table-hover mb-0">
              <thead><tr><th>Item</th><th>Container</th><th>Status</th><th></th></tr></thead>
              <tbody>
                {% for ex in exclusions %}
                <tr>
                  <td><strong>{{ ex.name }}</strong></td>
                  <td class="text-muted">{{ ex.container }}</td>
                  <td>
                    <form method="post" action="/exclusions/toggle/{{ ex.name | urlencode }}" class="d-inline">
                      {% if ex.in_stock %}
                      <button type="submit" class="btn btn-sm btn-outline-secondary">In Stock</button>
                      {% else %}
                      <button type="submit" class="btn btn-sm btn-danger">Out of Stock</button>
                      {% endif %}
                    </form>
                  </td>
                  <td class="text-end">
                    <a href="/exclusions/delete/{{ ex.name | urlencode }}" class="btn btn-sm btn-outline-danger"
                       onclick="return confirm('Remove {{ ex.name }} from exclusions?')">Del</a>
                  </td>
                </tr>
                {% endfor %}
                {% if not exclusions %}
                <tr><td colspan="4" class="text-center text-muted py-4">No exclusions yet.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        </div>
      </div>
      <div class="col-md-4">
        <div class="card">
          <div class="card-header">Add Exclusion</div>
          <div class="card-body">
            <p class="small text-muted mb-3">Staples you always have on hand. Recipes won't decrement these automatically.</p>
            <form method="post" action="/exclusions/add">
              <div class="mb-3">
                <label class="form-label fw-bold">Item Name</label>
                <input name="name" class="form-control" placeholder="e.g. Salt" required>
              </div>
              <div class="mb-3">
                <label class="form-label fw-bold">Container</label>
                <input name="container" class="form-control" placeholder="e.g. 1 shaker, 1 5lb bag">
              </div>
              <button type="submit" class="btn btn-primary w-100">Add</button>
            </form>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Backups Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'backups' %}show active{% endif %}"
       id="backups" role="tabpanel">
    <div class="row g-4">

      <div class="col-md-4">
        <div class="card mb-4">
          <div class="card-header">Backup Settings</div>
          <div class="card-body">
            <form method="post" action="/settings/backups" id="backup-cfg-form">

              <div class="mb-3">
                <label class="form-label fw-bold">When to back up</label>
                <div class="d-flex flex-column gap-2">
                  <label class="d-flex align-items-center gap-2" style="cursor:pointer">
                    <input type="radio" name="backup_mode" value="actions"
                           {% if backup_mode == 'actions' %}checked{% endif %}
                           onchange="updateBackupUI()">
                    Every X updates
                  </label>
                  <label class="d-flex align-items-center gap-2" style="cursor:pointer">
                    <input type="radio" name="backup_mode" value="interval"
                           {% if backup_mode == 'interval' %}checked{% endif %}
                           onchange="updateBackupUI()">
                    Every X hours
                  </label>
                  <label class="d-flex align-items-center gap-2" style="cursor:pointer">
                    <input type="radio" name="backup_mode" value="manual"
                           {% if backup_mode == 'manual' %}checked{% endif %}
                           onchange="updateBackupUI()">
                    Manual only
                  </label>
                </div>
              </div>

              <div id="cfg-actions" class="mb-3">
                <label class="form-label fw-bold">Updates between backups</label>
                <input type="number" name="backup_every_n" min="1" max="500"
                       class="form-control" value="{{ backup_every_n }}">
              </div>

              <div id="cfg-interval" class="mb-3">
                <label class="form-label fw-bold">Hours between backups</label>
                <input type="number" name="backup_interval_hrs" min="0.25" step="0.25"
                       class="form-control" value="{{ backup_interval }}">
                <div class="form-text">Use decimals for less than an hour (e.g. 0.5 = 30 min).</div>
              </div>

              <div class="mb-3">
                <label class="form-label fw-bold">Backups to keep</label>
                <input type="number" name="max_backups" min="1" max="50"
                       class="form-control" value="{{ backup_max }}">
              </div>

              <button type="submit" class="btn btn-primary w-100">Save</button>
            </form>
          </div>
        </div>
        <div class="card">
          <div class="card-header">Status &amp; Manual Backup</div>
          <div class="card-body">
            <p class="small text-muted mb-3">
              {{ backup_last_label }}<br>
              <span id="status-actions">{{ backup_count }} of {{ backup_every_n }} updates since last backup</span>
              <span id="status-interval">Backing up every {{ backup_interval }}h</span>
              <span id="status-manual">Manual mode — use Backup Now to create a backup</span>
            </p>
            <form method="post" action="/backups/now">
              <button type="submit" class="btn btn-outline-primary w-100">Backup Now</button>
            </form>
          </div>
        </div>
        <div class="card mt-4">
          <div class="card-header">Upload Backup</div>
          <div class="card-body">
            <p class="small text-muted mb-3">Upload an xlsx backup from your computer. It will appear in the list and can be restored from there.</p>
            <form method="post" action="/backups/upload" enctype="multipart/form-data">
              <div class="mb-3">
                <input type="file" name="backup_file" accept=".zip,.xlsx" class="form-control" required>
              </div>
              <button type="submit" class="btn btn-outline-secondary w-100">⬆ Upload</button>
            </form>
          </div>
        </div>
      </div>

<script>
function updateBackupUI() {
  const mode = document.querySelector('input[name=backup_mode]:checked').value;
  document.getElementById('cfg-actions').style.display   = mode === 'actions'  ? '' : 'none';
  document.getElementById('cfg-interval').style.display  = mode === 'interval' ? '' : 'none';
  document.getElementById('status-actions').style.display  = mode === 'actions'  ? '' : 'none';
  document.getElementById('status-interval').style.display = mode === 'interval' ? '' : 'none';
  document.getElementById('status-manual').style.display   = mode === 'manual'   ? '' : 'none';
}
updateBackupUI();
</script>

      <div class="col-md-8">
        <div class="card">
          <div class="card-header">
            Saved Backups
            <small class="fw-normal ms-2" style="font-size:.8em;opacity:.8">{{ backups|length }} file{{ 's' if backups|length != 1 else '' }}</small>
          </div>
          {% if not backups %}
          <div class="card-body text-muted">No backups yet — one will be created the next time you save an item.</div>
          {% else %}
          <div class="table-responsive">
            <table class="table table-hover mb-0">
              <thead><tr><th>Date</th><th>Size</th><th></th></tr></thead>
              <tbody>
                {% for b in backups %}
                <tr>
                  <td><strong>{{ b.modified }}</strong></td>
                  <td class="text-muted small">{{ b.size }}</td>
                  <td class="text-end">
                    <a href="/backups/download/{{ b.name }}" class="btn btn-sm btn-outline-secondary me-1">⬇ Download</a>
                    <form method="post" action="/backups/restore/{{ b.name }}" class="d-inline"
                          onsubmit="return confirm('Restore this backup? Your current inventory will be backed up first.')">
                      <button type="submit" class="btn btn-sm btn-outline-primary">↩ Restore</button>
                    </form>
                  </td>
                </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>
          {% endif %}
        </div>
      </div>

    </div>
  </div>

  <!-- ── Appearance Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'appearance' %}show active{% endif %}"
       id="appearance" role="tabpanel">
    <div class="row g-4">
      <div class="col-md-4">
        <div class="card">
          <div class="card-header">Theme Color</div>
          <div class="card-body">
            <p class="small text-muted mb-3">Pick the accent color used across the app — navbar, buttons, card headers, and table headers.</p>
            <form method="post" action="/settings/appearance">
              <div class="mb-3">
                <label class="form-label fw-bold">Accent Color</label>
                <div class="d-flex align-items-center gap-3">
                  <input type="color" name="accent_color" value="{{ accent_color }}"
                         class="form-control form-control-color" style="width:60px;height:40px;padding:2px;">
                  <span class="font-monospace small text-muted" id="color-hex">{{ accent_color }}</span>
                </div>
              </div>
              <div class="mb-3">
                <label class="form-label fw-bold small">Presets</label>
                <div class="d-flex flex-wrap gap-2">
                  {% for label, hex in [('Purple', '#6B2D8B'), ('Blue', '#1A56DB'), ('Green', '#057A55'), ('Red', '#C81E1E'), ('Orange', '#B45309'), ('Teal', '#0E7490'), ('Pink', '#9D174D')] %}
                  <button type="button" class="btn btn-sm"
                          style="background:{{ hex }};border-color:{{ hex }};color:white;min-width:60px;"
                          onclick="document.querySelector('input[name=accent_color]').value='{{ hex }}';document.getElementById('color-hex').textContent='{{ hex }}';">
                    {{ label }}
                  </button>
                  {% endfor %}
                </div>
              </div>
              <button type="submit" class="btn btn-primary w-100">Apply</button>
            </form>
          </div>
        </div>
      </div>
      <div class="col-md-4">
        <div class="card">
          <div class="card-header">Preview</div>
          <div class="card-body">
            <p class="small text-muted mb-3">A live preview will appear after you click Apply.</p>
            <div class="d-flex flex-column gap-2">
              <button class="btn btn-primary">Primary Button</button>
              <button class="btn btn-outline-primary">Outline Button</button>
              <div class="p-2 rounded text-white text-center small" style="background:var(--purple);">Navbar / Card Header</div>
              <div class="p-2 rounded text-center small" style="background:var(--light-purple);color:var(--dark-purple);">Light Accent</div>
            </div>
          </div>
        </div>
      </div>
    </div>
<script>
document.querySelector('input[name=accent_color]').addEventListener('input', function() {
  document.getElementById('color-hex').textContent = this.value;
});
</script>
  </div>

  <!-- ── Discord Bot Tab ── -->
  <div class="tab-pane fade {% if active_tab == 'discord' %}show active{% endif %}"
       id="discord" role="tabpanel">
    <div class="row g-4">
      <div class="col-md-6">
        <div class="card mb-4">
          <div class="card-header">Discord Bot</div>
          <div class="card-body">
            <p class="text-muted small mb-3">
              Create a bot at <strong>discord.com/developers/applications</strong>, copy the token, and paste it here.
              The bot will connect to any server it has been invited to.
            </p>
            <div class="mb-2 d-flex align-items-center gap-2">
              <span class="fw-bold small">Status:</span>
              <span id="bot-status-badge" class="badge bg-secondary">Checking...</span>
            </div>
            <form method="post" action="/settings/discord" class="mt-3">
              <div class="mb-3">
                <label class="form-label fw-bold">Bot Token</label>
                <input type="password" name="token" class="form-control font-monospace"
                       placeholder="{{ 'Paste a new token to replace the saved one' if token_set else 'Paste your Discord bot token here' }}">
                {% if token_set %}
                <div class="form-text text-success">Token is saved — leave blank to keep it</div>
                {% else %}
                <div class="form-text text-danger">No token set — bot is offline</div>
                {% endif %}
                <div class="form-text text-muted">Your token is never included in backups.</div>
              </div>
              <div class="mb-3">
                <label class="form-label fw-bold">Low-Stock Alert Channel ID</label>
                <input type="text" name="alert_channel" class="form-control font-monospace"
                       placeholder="Discord channel ID (right-click channel → Copy ID)"
                       value="{{ alert_channel }}">
                <div class="form-text">When set, the bot will post a warning here whenever an item drops below its minimum.</div>
              </div>
              <div class="d-flex gap-2 flex-wrap">
                <button type="submit" name="action" value="save" class="btn btn-primary">Save Settings</button>
                <button type="submit" name="action" value="save_restart" class="btn btn-success">Save &amp; Restart Bot</button>
                {% if token_set %}
                <button type="submit" name="action" value="restart" class="btn btn-outline-primary">Restart Bot</button>
                <button type="submit" name="action" value="stop" class="btn btn-outline-secondary">Stop Bot</button>
                <button type="submit" name="action" value="remove" class="btn btn-outline-danger ms-auto"
                        onclick="return confirm('Remove the Discord token and disconnect the bot?')">Remove Connection</button>
                {% endif %}
              </div>
            </form>
          </div>
        </div>
      </div>
      <div class="col-md-6">
        <div class="card mb-4">
          <div class="card-header">How to Set Up Your Discord Bot</div>
          <div class="card-body small text-muted">
            <ol class="ps-3 mb-0">
              <li class="mb-2">Go to <strong>discord.com/developers/applications</strong> and click <strong>New Application</strong></li>
              <li class="mb-2">Go to the <strong>Bot</strong> tab → click <strong>Add Bot</strong></li>
              <li class="mb-2">Under <strong>Privileged Gateway Intents</strong>, enable <strong>Message Content Intent</strong></li>
              <li class="mb-2">Click <strong>Reset Token</strong> → copy the token → paste it above</li>
              <li class="mb-2">Go to <strong>OAuth2 → URL Generator</strong>, select <strong>bot</strong> scope and <strong>Send Messages / Read Message History</strong> permissions</li>
              <li class="mb-2">Use the generated URL to invite the bot to your server</li>
              <li>For low-stock alerts: right-click the target channel → <strong>Copy Channel ID</strong> (enable Developer Mode in Discord settings first)</li>
            </ol>
          </div>
        </div>
        <div class="card">
          <div class="card-header d-flex justify-content-between align-items-center">
            Bot Log
            <button type="button" class="btn btn-outline-secondary btn-sm" onclick="loadBotLog()">Refresh</button>
          </div>
          <div class="card-body p-0">
            <pre id="bot-log" class="mb-0 p-3 font-monospace small"
                 style="max-height:220px;overflow-y:auto;background:#1a1a1a;color:#d4f0d4;border-radius:0 0 10px 10px;white-space:pre-wrap;word-break:break-all;">Loading…</pre>
          </div>
        </div>
      </div>
    </div>
  </div>

</div>

<script>
fetch('/settings/bot-status')
  .then(r => r.json())
  .then(d => {
    const b = document.getElementById('bot-status-badge');
    b.className = d.running ? 'badge bg-success' : 'badge bg-danger';
    b.textContent = d.running ? 'Online' : 'Offline';
  });

function loadBotLog() {
  fetch('/settings/bot-log')
    .then(r => r.json())
    .then(d => {
      const el = document.getElementById('bot-log');
      el.textContent = d.log || '(no log yet)';
      el.scrollTop = el.scrollHeight;
    });
}
loadBotLog();

document.querySelectorAll('.qty-display').forEach(span => {
  span.addEventListener('click', () => {
    span.classList.add('d-none');
    const form = span.nextElementSibling;
    form.classList.remove('d-none');
    form.style.display = 'inline';
    form.querySelector('input[type=number]').focus();
    form.querySelector('input[type=number]').select();
  });
});
document.querySelectorAll('.qty-cancel').forEach(btn => {
  btn.addEventListener('click', () => {
    const form = btn.closest('form');
    form.classList.add('d-none');
    form.previousElementSibling.classList.remove('d-none');
  });
});
</script>
<script src="https://cdn.jsdelivr.net/npm/sortablejs@1.15.2/Sortable.min.js"></script>
<script>
(function() {
  var locBody = document.getElementById('locations-sortable');
  if (locBody) {
    Sortable.create(locBody, {
      handle: 'td:first-child',
      animation: 150,
      onEnd: function() {
        var order = Array.from(locBody.querySelectorAll('tr[data-name]'))
                        .map(function(r){ return r.dataset.name; });
        fetch('/locations/reorder', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({order: order})
        });
      }
    });
  }

  var unitBody = document.getElementById('units-sortable');
  if (unitBody) {
    Sortable.create(unitBody, {
      handle: 'td:first-child',
      animation: 150,
      onEnd: function() {
        var order = Array.from(unitBody.querySelectorAll('tr[data-unit]'))
                        .map(function(r){ return r.dataset.unit; });
        fetch('/units/reorder', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({order: order})
        });
      }
    });
  }
})();
</script>
""")

RECIPE_DETAIL_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<div class="mb-3 d-flex gap-2 align-items-center flex-wrap">
  <a href="/recipes" class="btn btn-outline-secondary btn-sm">← Back to Recipes</a>
  <a href="/recipe/edit/{{ slug }}" class="btn btn-outline-primary btn-sm">Edit Recipe</a>
  <form method="post" action="/recipe/{{ slug }}/use" class="ms-auto"
        onsubmit="return confirm('Remove all ingredient quantities from your pantry inventory?')">
    <button type="submit" class="btn btn-warning btn-sm">🍳 Use Recipe (Remove from Pantry)</button>
  </form>
  <a href="/recipe/delete/{{ slug }}" class="btn btn-sm btn-outline-danger"
     onclick="return confirm('Delete {{ recipe.name }}?')">Delete</a>
</div>

{% if used_count is defined %}
<div class="alert mb-3" style="background:#e8f5e9;border:1px solid #a5d6a7;color:#2e7d32;">
  ✅ Recipe used — <strong>{{ used_count }}</strong> ingredient(s) removed from pantry.
  {% if not_found_count %}<span class="text-muted"> ({{ not_found_count }} item(s) not found in pantry.)</span>{% endif %}
</div>
{% endif %}
<div class="card mb-4">
  <div class="card-header">{{ recipe.name }}</div>
  <div class="card-body">
    <div class="row mb-3 align-items-center">
      <div class="col text-muted small">
        <strong>{{ recipe.meal_type }}</strong> &nbsp;·&nbsp;
        Prep {{ recipe.prep_time }}{% if recipe.cook_time %} &nbsp;·&nbsp; Cook {{ recipe.cook_time }}{% endif %}
      </div>
      <div class="col-auto d-flex align-items-center gap-2">
        <label class="form-label mb-0 small fw-bold">Servings:</label>
        <input id="servings-input" type="number" min="1" value="{{ recipe.servings }}"
               class="form-control form-control-sm" style="width:70px;"
               oninput="scaleRecipe(this.value)">
        <span class="text-muted small">(original: {{ recipe.servings }})</span>
      </div>
    </div>

    {% set m = recipe.macros_per_serving %}
    {% if m and m.calories %}
    <div class="row g-2 mb-4" id="macro-tiles">
      <div class="col-6 col-md-3">
        <div class="card text-center" style="border-color:#E8D5F0;">
          <div class="card-body py-2">
            <div class="fw-bold macro-val" data-base="{{ m.calories }}" style="color:var(--purple);font-size:1.3em;">{{ m.calories }}</div>
            <div class="text-muted small">Cal / serving</div>
          </div>
        </div>
      </div>
      <div class="col-6 col-md-3">
        <div class="card text-center" style="border-color:#E8D5F0;">
          <div class="card-body py-2">
            <div class="fw-bold macro-val" data-base="{{ m.protein_g }}" style="color:var(--purple);font-size:1.3em;">{{ m.protein_g }}g</div>
            <div class="text-muted small">Protein / serving</div>
          </div>
        </div>
      </div>
      <div class="col-6 col-md-3">
        <div class="card text-center" style="border-color:#E8D5F0;">
          <div class="card-body py-2">
            <div class="fw-bold macro-val" data-base="{{ m.carbs_g }}" style="color:var(--purple);font-size:1.3em;">{{ m.carbs_g }}g</div>
            <div class="text-muted small">Carbs / serving</div>
          </div>
        </div>
      </div>
      <div class="col-6 col-md-3">
        <div class="card text-center" style="border-color:#E8D5F0;">
          <div class="card-body py-2">
            <div class="fw-bold macro-val" data-base="{{ m.fat_g }}" style="color:var(--purple);font-size:1.3em;">{{ m.fat_g }}g</div>
            <div class="text-muted small">Fat / serving</div>
          </div>
        </div>
      </div>
    </div>
    {% endif %}

    <div class="row g-4">
      <div class="col-md-5">
        <h5 style="color:#4A1E6B;">Ingredients</h5>
        <ul class="list-unstyled mb-0" id="ingredients-list">
          {% for ing in recipe.ingredients %}
          <li class="py-1 ing-row" style="border-bottom:1px solid #f0e6f8;"
              data-amount="{{ ing.amount }}">
            {% if have is defined and ing.item|lower in have %}✅{% else %}•{% endif %}
            <strong class="ing-amount">{{ ing.amount }}</strong> {{ ing.item }}
          </li>
          {% endfor %}
        </ul>
      </div>
      <div class="col-md-7">
        <h5 style="color:#4A1E6B;">Instructions</h5>
        <ol class="ps-3">
          {% for step in recipe.instructions %}
          <li class="mb-2">{{ step }}</li>
          {% endfor %}
        </ol>
      </div>
    </div>

    {% if recipe.substitutions %}
    <hr style="border-color:#E8D5F0;">
    <h5 style="color:#4A1E6B;">Substitutions</h5>
    <div class="row g-2">
      {% for key, val in recipe.substitutions.items() %}
      <div class="col-md-6">
        <span class="fw-bold text-muted small">{{ key.replace('_', ' ').title() }}:</span>
        <span class="small"> {{ val }}</span>
      </div>
      {% endfor %}
    </div>
    {% endif %}

    {% if recipe.notes %}
    <hr style="border-color:#E8D5F0;">
    <p class="mb-0"><strong>Notes:</strong> {{ recipe.notes }}</p>
    {% endif %}
  </div>
</div>

<script>
const origServings = {{ recipe.servings }};

function scaleRecipe(newServings) {
  const n = parseFloat(newServings);
  if (!n || n <= 0) return;
  const ratio = n / origServings;

  // Scale ingredient amounts
  document.querySelectorAll('.ing-row').forEach(li => {
    const origAmount = li.dataset.amount || '';
    const scaled = scaleAmount(origAmount, ratio);
    li.querySelector('.ing-amount').textContent = scaled;
  });

  // Scale macros per serving (keep them per-serving — they don't change with serving count)
  // Instead show total if scaled up
}

function scaleAmount(text, ratio) {
  if (!text) return text;
  // Find the leading numeric portion (digits, spaces, slashes, dots, vulgar fractions)
  const numChars = new Set('0123456789 ./\t');
  const vulgar   = new Set([...'½¼¾⅓⅔⅛⅜⅝⅞⅐⅑⅒⅕⅖⅗⅘⅙⅚']);
  let i = 0;
  while (i < text.length && (numChars.has(text[i]) || vulgar.has(text[i]))) i++;
  if (i === 0) return text;
  const match = text.slice(0, i);
  const rest  = text.slice(i);
  const num   = parseFraction(match.trim());
  if (isNaN(num) || num === 0) return text;
  return formatNum(num * ratio) + (rest ? ' ' + rest.trimStart() : '');
}

function parseFraction(s) {
  const fracs = {'½':0.5,'¼':0.25,'¾':0.75,'⅓':0.333,'⅔':0.667,'⅛':0.125,'⅜':0.375,'⅝':0.625,'⅞':0.875};
  if (fracs[s]) return fracs[s];
  // "1 1/2" or "1/2"
  const parts = s.split(/ +/);
  let val = 0;
  for (const p of parts) {
    if (fracs[p]) { val += fracs[p]; continue; }
    if (p.includes('/')) {
      const [a,b] = p.split('/');
      val += parseFloat(a) / parseFloat(b);
    } else {
      val += parseFloat(p) || 0;
    }
  }
  return val;
}

function formatNum(n) {
  if (n === Math.floor(n)) return String(Math.floor(n));
  // Common fractions
  const eighths = Math.round(n * 8) / 8;
  const map = {0.125:'⅛',0.25:'¼',0.375:'⅜',0.5:'½',0.625:'⅝',0.75:'¾',0.875:'⅞'};
  const whole = Math.floor(eighths);
  const frac  = Math.round((eighths - whole) * 8) / 8;
  const fracStr = map[frac] || '';
  if (whole === 0) return fracStr || n.toFixed(2).replace(/[.]?0+$/, '');
  return fracStr ? `${whole} ${fracStr}` : n.toFixed(2).replace(/[.]?0+$/, '');
}
</script>
""")



SHOPPING_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<div class="d-flex justify-content-between align-items-center mb-3">
  <h1>Shopping List</h1>
</div>
<div class="row g-4">
  <div class="col-md-6">
    <div class="card mb-4">
      <div class="card-header">🛒 Restock Needs <small class="fw-normal ms-2" style="font-size:.8em;opacity:.8">items below minimum</small></div>
      {% if not low %}
      <div class="card-body text-muted">Nothing below minimum — pantry is stocked.</div>
      {% else %}
      <div class="table-responsive">
        <table class="table table-hover mb-0">
          <thead><tr><th>Item</th><th>Have</th><th>Need</th><th>Unit</th></tr></thead>
          <tbody>
            {% for item, current, minimum, unit in low %}
            <tr>
              <td><strong>{{ item }}</strong></td>
              <td>{{ current }}</td>
              <td class="fw-bold text-danger">{{ (minimum - current)|round(1) }}</td>
              <td class="text-muted">{{ unit | pluralize_unit(current) }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      {% endif %}
    </div>
  </div>
  <div class="col-md-6">
    <div class="card">
      <div class="card-header">🍳 Missing Recipe Ingredients</div>
      <div class="card-body">
        <form method="get" id="recipe-form" class="mb-3">
          <select name="recipe" class="form-select search-bar"
                  onchange="this.form.submit()">
            <option value="">— Pick a recipe —</option>
            {% for r in all_recipes %}
            <option value="{{ r.slug }}" {% if selected_recipe == r.slug %}selected{% endif %}>{{ r.name }}</option>
            {% endfor %}
          </select>
        </form>
        {% if missing_ingredients %}
        <ul class="list-unstyled mb-2">
          {% for ing in missing_ingredients %}
          <li class="py-1" style="border-bottom:1px solid #f0e6f8;">
            ❌ <strong>{{ ing.amount }}</strong> {{ ing.item }}
          </li>
          {% endfor %}
        </ul>
        <form method="post" action="/shopping/list/add-recipe">
          <input type="hidden" name="recipe" value="{{ selected_recipe }}">
          <button type="submit" class="btn btn-primary btn-sm w-100">
            ➕ Add {{ missing_ingredients|length }} missing item{{ 's' if missing_ingredients|length != 1 else '' }} to shopping list
          </button>
        </form>
        {% elif selected_recipe %}
        <div class="text-success">✅ You have everything for this recipe!</div>
        {% else %}
        <p class="text-muted small mb-0">Pick a recipe above to see what you're missing.</p>
        {% endif %}
      </div>
    </div>
  </div>
</div>

{% if out_of_stock_exclusions %}
<div class="card mt-4">
  <div class="card-header">📦 Out of Stock (Staples)</div>
  <div class="table-responsive">
    <table class="table table-hover mb-0">
      <thead><tr><th>Item</th><th>Container</th><th></th></tr></thead>
      <tbody>
        {% for ex in out_of_stock_exclusions %}
        <tr>
          <td><strong>{{ ex.name }}</strong></td>
          <td class="text-muted">{{ ex.container }}</td>
          <td class="text-end">
            <form method="post" action="/exclusions/toggle/{{ ex.name | urlencode }}" class="d-inline">
              <button type="submit" class="btn btn-sm btn-outline-success">Mark In Stock</button>
            </form>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<div class="card mt-4">
  <div class="card-header d-flex justify-content-between align-items-center">
    <span>🛍 Shopping List
      {% if shopping_list %}
      <small class="fw-normal ms-2" style="font-size:.8em;opacity:.8">{{ shopping_list|length }} item{{ 's' if shopping_list|length != 1 else '' }}</small>
      {% endif %}
    </span>
    {% if shopping_list %}
    <div class="d-flex gap-2">
      <form method="post" action="/shopping/list/clear-checked" class="d-inline">
        <button type="submit" class="btn btn-outline-secondary btn-sm">Remove checked</button>
      </form>
      <form method="post" action="/shopping/list/clear-all" class="d-inline"
            onsubmit="return confirm('Clear the entire shopping list?')">
        <button type="submit" class="btn btn-outline-danger btn-sm">Clear all</button>
      </form>
    </div>
    {% endif %}
  </div>
  {% if not shopping_list %}
  <div class="card-body text-muted">
    No items yet — pick a recipe above and add its missing ingredients.
  </div>
  {% else %}
  <div class="table-responsive">
    <table class="table table-hover mb-0">
      <thead><tr><th style="width:2rem"></th><th>Item</th><th>Amount</th><th>From Recipe</th><th></th></tr></thead>
      <tbody>
        {% for item in shopping_list %}
        <tr class="{{ 'text-muted' if item.checked else '' }}" style="{{ 'opacity:.5;' if item.checked else '' }}">
          <td>
            <form method="post" action="/shopping/list/toggle/{{ item.id }}">
              <button type="submit" class="btn btn-sm p-0 border-0 bg-transparent"
                      title="{{ 'Uncheck' if item.checked else 'Check off' }}">
                {{ '☑' if item.checked else '☐' }}
              </button>
            </form>
          </td>
          <td><strong>{{ item.name }}</strong></td>
          <td>{{ item.amount or '—' }}</td>
          <td class="text-muted small">{{ item.note or '' }}</td>
          <td>
            <a href="/shopping/list/remove/{{ item.id }}"
               class="btn btn-sm btn-outline-danger"
               onclick="return confirm('Remove {{ item.name }}?')">✕</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endif %}
</div>
""")

BULK_ADD_HTML = BASE_HTML.replace('{% block content %}{% endblock %}', """
<h1>Bulk Add Items</h1>
<p class="text-muted">Add multiple items at once. One item per line.</p>
<div class="card mt-3" style="max-width:640px">
  <div class="card-header">Paste Items</div>
  <div class="card-body">
    <p class="small text-muted mb-2">
      Format each line as: <code>Item Name, Quantity, Unit, Location</code><br>
      Quantity, Unit, and Location are optional. Examples:<br>
      <code>Black Beans, 3, 15oz can, Pantry</code><br>
      <code>Olive Oil, 1, bottle</code><br>
      <code>Salt</code>
    </p>
    <form method="post">
      <div class="mb-3">
        <textarea name="bulk" class="form-control font-monospace" rows="12"
                  placeholder="Black Beans, 3, 15oz can, Pantry&#10;Olive Oil, 1, bottle&#10;Salt"></textarea>
      </div>
      <div class="mb-3">
        <label class="form-label fw-bold">Default Location (if not specified per line)</label>
        <select name="default_location" class="form-select">
          {% for l in locations %}
          <option>{{ l }}</option>
          {% endfor %}
        </select>
      </div>
      <div class="d-flex gap-2">
        <button type="submit" class="btn btn-primary px-4">Add All</button>
        <a href="/" class="btn btn-outline-secondary">Cancel</a>
      </div>
    </form>
    {% if results %}
    <hr>
    <h6>Results:</h6>
    <ul class="small">
      {% for msg in results %}
      <li>{{ msg }}</li>
      {% endfor %}
    </ul>
    {% endif %}
  </div>
</div>
""")


@app.route('/')
def index():
    tab = request.args.get('tab', 'items')
    q   = request.args.get('q', '').lower()
    loc = request.args.get('loc', '')
    wb        = load_wb()
    all_raw   = get_all_rows(wb)
    locations = get_location_sheets()
    rows = [(f'{s}:{r}', d) for s, r, d in all_raw]
    if q:
        rows = [(i, r) for i, r in rows if q in str(r.get('Item', '')).lower()]
    if loc:
        rows = [(i, r) for i, r in rows if str(r.get('Location', '')) == loc or i.startswith(loc + ':')]
    return render_template_string(INDEX_HTML, items=rows, total=len(rows),
                                  locations=locations, q=q, loc=loc,
                                  tab=tab, bulk_results=[], units=load_units())


@app.route('/add', methods=['GET', 'POST'])
def add():
    location_sheets = get_location_sheets()
    if request.method == 'POST':
        item     = to_title((request.form.get('Item', '') or '').strip())
        quantity = request.form.get('Quantity', '') or None
        unit     = request.form.get('Unit', '') or None
        location = request.form.get('Location', '')
        sheet_name = location if location in location_sheets else location_sheets[0]
        wb  = load_wb()
        ws  = wb[sheet_name]
        # Check if same item + unit already exists in this sheet — merge if so
        existing = next((r for r in range(2, ws.max_row + 1)
                         if str(ws.cell(row=r, column=1).value or '').lower() == (item or '').lower()
                         and str(ws.cell(row=r, column=3).value or '').lower() == (unit or '').lower()), None)
        if existing:
            try:
                cur = float(str(ws.cell(row=existing, column=2).value or 0))
                add_qty = float(quantity) if quantity else 1
                merged = cur + add_qty
                ws.cell(row=existing, column=2).value = int(merged) if merged == int(merged) else merged
            except (ValueError, TypeError):
                pass
        else:
            row = next((r for r in range(2, ws.max_row + 2) if not ws.cell(row=r, column=1).value), ws.max_row + 1)
            ws.cell(row=row, column=1).value = item or None
            ws.cell(row=row, column=2).value = quantity
            ws.cell(row=row, column=3).value = unit
            ws.cell(row=row, column=4).value = sheet_name
        save_wb(wb)
        return redirect(url_for('index', tab='add'))
    return redirect(url_for('index', tab='add'))


@app.route('/edit/<path:row_key>', methods=['GET', 'POST'])
def edit(row_key):
    sheet_name, row_idx = row_key.rsplit(':', 1)
    row_idx = int(row_idx)
    location_sheets = get_location_sheets()
    wb = load_wb()
    ws = wb[sheet_name]
    if request.method == 'POST':
        item     = to_title((request.form.get('Item', '') or '').strip())
        quantity = request.form.get('Quantity', '') or None
        unit     = request.form.get('Unit', '') or None
        location = request.form.get('Location', '') or sheet_name
        new_sheet = location if location in location_sheets else sheet_name
        if new_sheet != sheet_name:
            # Clear old row
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).value = None
            ws2 = wb[new_sheet]
            # Check if same item + unit already exists in destination — merge quantities if so
            existing = next((r for r in range(2, ws2.max_row + 1)
                             if str(ws2.cell(row=r, column=1).value or '').lower() == (item or '').lower()
                             and str(ws2.cell(row=r, column=3).value or '').lower() == (unit or '').lower()), None)
            if existing:
                try:
                    cur = float(str(ws2.cell(row=existing, column=2).value or 0))
                    add_qty = float(quantity) if quantity else 0
                    merged = cur + add_qty
                    ws2.cell(row=existing, column=2).value = int(merged) if merged == int(merged) else merged
                except (ValueError, TypeError):
                    pass
            else:
                new_row = next((r for r in range(2, ws2.max_row + 2) if not ws2.cell(row=r, column=1).value), ws2.max_row + 1)
                ws2.cell(row=new_row, column=1).value = item or None
                ws2.cell(row=new_row, column=2).value = quantity
                ws2.cell(row=new_row, column=3).value = unit
                ws2.cell(row=new_row, column=4).value = new_sheet
        else:
            # Same sheet — check for collision with another row (different row_idx, same name + unit)
            collision = next((r for r in range(2, ws.max_row + 1)
                              if r != row_idx
                              and str(ws.cell(row=r, column=1).value or '').lower() == (item or '').lower()
                              and str(ws.cell(row=r, column=3).value or '').lower() == (unit or '').lower()), None)
            if collision:
                try:
                    cur = float(str(ws.cell(row=collision, column=2).value or 0))
                    add_qty = float(quantity) if quantity else 0
                    merged = cur + add_qty
                    ws.cell(row=collision, column=2).value = int(merged) if merged == int(merged) else merged
                except (ValueError, TypeError):
                    pass
                # Clear the row being edited since we merged into the existing one
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).value = None
            else:
                ws.cell(row=row_idx, column=1).value = item or None
                ws.cell(row=row_idx, column=2).value = quantity
                ws.cell(row=row_idx, column=3).value = unit
                ws.cell(row=row_idx, column=4).value = sheet_name
        save_wb(wb)
        return redirect(url_for('index'))
    values = {field: ws.cell(row=row_idx, column=col).value or ''
              for col, field in enumerate(COLS, 1)}
    return render_template_string(EDIT_HTML, title=f'Edit: {values.get("Item", "")}',
                                  values=values, locations=location_sheets, units=load_units())


@app.route('/delete/<path:row_key>')
def delete(row_key):
    sheet_name, row_idx = row_key.rsplit(':', 1)
    row_idx = int(row_idx)
    wb = load_wb()
    ws = wb[sheet_name]
    for col in range(1, 9):
        ws.cell(row=row_idx, column=col).value = None
    save_wb(wb)
    return redirect(url_for('index'))


@app.route('/can-make')
def can_make():
    return redirect(url_for('recipes', tab='canmake'))


@app.route('/minimums')
def minimums():
    return redirect(url_for('settings', tab='minimums'))


@app.route('/minimums/set', methods=['POST'])
def minimums_set():
    item = to_title((request.form.get('item', '') or '').strip())
    qty_raw = request.form.get('qty', '0')
    try:
        qty = float(qty_raw)
        qty = int(qty) if qty == int(qty) else qty
    except ValueError:
        qty = 0
    if item and qty > 0:
        set_minimum(item, qty)
    return redirect(url_for('settings', tab='minimums'))


@app.route('/minimums/delete/<item>')
def minimums_delete(item):
    delete_minimum(item)
    return redirect(url_for('settings', tab='minimums'))


@app.route('/recipes')
def recipes():
    tab        = request.args.get('tab', 'library')
    q          = request.args.get('q', '').lower()
    meal       = request.args.get('meal', '')
    cm_meal    = request.args.get('cm_meal', '')
    ingredient = request.args.get('ingredient', '').strip().lower()

    all_recipes = []
    if DATA_DIR.exists():
        for path in sorted(DATA_DIR.glob('*.json')):
            try:
                r = json.load(open(path, encoding='utf-8'))
            except Exception:
                continue
            r['slug'] = path.stem
            all_recipes.append(r)

    lib_recipes = [r for r in all_recipes
                   if (not q or q in r.get('name', '').lower())
                   and (not meal or meal.lower() in r.get('meal_type', '').lower())]
    lib_recipes.sort(key=lambda x: x.get('name', ''))

    cm_recipes = []
    if tab == 'canmake':
        wb   = load_wb()
        have = {str(d.get('Item', '')).lower() for _, _, d in get_all_rows(wb) if d.get('Item')}
        for r in all_recipes:
            if cm_meal and cm_meal.lower() not in r.get('meal_type', '').lower():
                continue
            ingredients = r.get('ingredients', [])
            needed = [(i.get('item', '') if isinstance(i, dict) else str(i)).lower() for i in ingredients]
            if not needed:
                continue
            if ingredient and not any(ingredient in n or n in ingredient for n in needed):
                continue
            have_count = sum(1 for n in needed if any(n in h or h in n for h in have))
            total = len(needed)
            pct   = int(have_count / total * 100) if total else 0
            missing = [n for n in needed if not any(n in h or h in n for h in have)]
            if pct == 100:
                status_class, status_label = 'text-success fw-bold', '✓ You have everything!'
            elif pct >= 60:
                status_class, status_label = 'text-warning fw-bold', f'~{pct}% — almost there'
            elif pct > 0:
                status_class, status_label = 'text-danger', f'{pct}% ({have_count}/{total} ingredients)'
            else:
                status_class, status_label = 'text-danger', '0% — missing everything'
            cm_recipes.append({'name': r.get('name', 'Untitled'), 'meal_type': r.get('meal_type', ''),
                               'prep_time': r.get('prep_time', ''), 'status_class': status_class,
                               'status_label': status_label, 'missing': missing[:6], 'pct': pct})
        cm_recipes.sort(key=lambda x: x['pct'], reverse=True)

    return render_template_string(RECIPES_HTML, recipes=lib_recipes, q=q, meal=meal,
                                  tab=tab, cm_recipes=cm_recipes, cm_meal=cm_meal,
                                  ingredient=ingredient, meal_types=MEAL_TYPES)


@app.route('/recipe/import')
def recipe_import():
    import requests as req
    from bs4 import BeautifulSoup

    url = request.args.get('url', '').strip()
    if not url:
        return json.dumps({'error': 'No URL provided'}), 400, {'Content-Type': 'application/json'}

    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        resp = req.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        return json.dumps({'error': f'Could not fetch page: {e}'}), 400, {'Content-Type': 'application/json'}

    soup = BeautifulSoup(resp.text, 'html.parser')

    # Find schema.org Recipe in any JSON-LD block
    recipe_data = None
    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string or '{}')
        except Exception:
            continue
        candidates = data if isinstance(data, list) else data.get('@graph', [data])
        for item in candidates:
            if isinstance(item, dict) and 'Recipe' in str(item.get('@type', '')):
                recipe_data = item
                break
        if recipe_data:
            break

    if not recipe_data:
        return json.dumps({'error': 'No recipe data found on this page. Try the original source URL instead of a saved CopyMeThat link.'}), 404, {'Content-Type': 'application/json'}

    def parse_duration(iso):
        if not iso:
            return ''
        h = re.search(r'(\d+)H', str(iso))
        m = re.search(r'(\d+)M', str(iso))
        parts = []
        if h:
            n = int(h.group(1))
            parts.append(f"{n} hour{'s' if n != 1 else ''}")
        if m:
            n = int(m.group(1))
            parts.append(f"{n} minute{'s' if n != 1 else ''}")
        return ' '.join(parts) or str(iso)

    def parse_servings(val):
        if not val:
            return 4
        if isinstance(val, list):
            val = val[0]
        m = re.search(r'\d+', str(val))
        return int(m.group()) if m else 4

    def extract_num(val):
        if not val:
            return 0
        m = re.search(r'[\d.]+', str(val))
        return float(m.group()) if m else 0

    UNITS = {
        'cup','cups','tablespoon','tablespoons','tbsp','teaspoon','teaspoons','tsp',
        'pound','pounds','lb','lbs','ounce','ounces','oz','gram','grams','g',
        'kilogram','kilograms','kg','liter','liters','ml','milliliter','milliliters',
        'can','cans','jar','jars','package','packages','pkg','clove','cloves',
        'slice','slices','piece','pieces','bunch','bunches','pinch','dash',
        'quart','quarts','pint','pints','gallon','gallons','sprig','sprigs',
    }

    def parse_ingredient(text):
        text = text.strip()
        parts = text.split()
        if not parts:
            return '', text
        # Must start with a digit or vulgar fraction to have an amount
        if not (parts[0][0].isdigit() or parts[0][0] in '½¼¾⅓⅔⅛⅜⅝⅞'):
            return '', text
        # Take leading number tokens (handles "1 1/2")
        i = 0
        while i < len(parts) and (parts[i].replace('/', '').replace('.', '').isdigit()
                                   or parts[i] in '½¼¾⅓⅔⅛⅜⅝⅞'):
            i += 1
        # Optionally include a unit word
        if i < len(parts) and parts[i].lower().rstrip('s.') in UNITS or (i < len(parts) and parts[i].lower() in UNITS):
            i += 1
        amount = ' '.join(parts[:i])
        item   = ' '.join(parts[i:])
        return amount, item

    def parse_instructions(raw):
        steps = []
        for item in (raw if isinstance(raw, list) else []):
            if isinstance(item, str):
                steps.append(item.strip())
            elif isinstance(item, dict):
                t = item.get('@type', '')
                if 'HowToStep' in t:
                    steps.append(item.get('text', '').strip())
                elif 'HowToSection' in t:
                    for sub in item.get('itemListElement', []):
                        if isinstance(sub, dict):
                            steps.append(sub.get('text', '').strip())
        return [s for s in steps if s]

    category = recipe_data.get('recipeCategory', '') or ''
    if isinstance(category, list):
        category = category[0] if category else ''
    meal_type = 'Dinner'
    for mt in MEAL_TYPES:
        if mt.lower() in str(category).lower():
            meal_type = mt
            break

    nutrition = recipe_data.get('nutrition') or {}
    ingredients = []
    for ing in recipe_data.get('recipeIngredient', []):
        amount, item = parse_ingredient(str(ing))
        ingredients.append({'amount': amount, 'item': item})

    result = {
        'name':         recipe_data.get('name', '').strip(),
        'meal_type':    meal_type,
        'servings':     parse_servings(recipe_data.get('recipeYield')),
        'prep_time':    parse_duration(recipe_data.get('prepTime')),
        'cook_time':    parse_duration(recipe_data.get('cookTime')),
        'calories':     int(extract_num(nutrition.get('calories'))),
        'protein':      extract_num(nutrition.get('proteinContent')),
        'carbs':        extract_num(nutrition.get('carbohydrateContent')),
        'fat':          extract_num(nutrition.get('fatContent')),
        'ingredients':  ingredients,
        'instructions': parse_instructions(recipe_data.get('recipeInstructions', [])),
        'notes':        (recipe_data.get('description') or '').strip(),
    }
    return json.dumps(result), 200, {'Content-Type': 'application/json'}


@app.route('/recipe/add', methods=['GET', 'POST'])
def recipe_add():
    if request.method == 'POST':
        name      = request.form.get('name', '').strip()
        meal_type = request.form.get('meal_type', '')
        servings  = int(request.form.get('servings') or 1)
        prep_time = request.form.get('prep_time', '').strip()
        cook_time = request.form.get('cook_time', '').strip()

        try:
            calories = float(request.form.get('calories') or 0)
            protein  = float(request.form.get('protein') or 0)
            carbs    = float(request.form.get('carbs') or 0)
            fat      = float(request.form.get('fat') or 0)
        except ValueError:
            calories = protein = carbs = fat = 0

        amounts     = request.form.getlist('amount[]')
        ingredients_raw = request.form.getlist('ingredient[]')
        ingredients = [{'amount': a.strip(), 'item': i.strip()}
                       for a, i in zip(amounts, ingredients_raw) if i.strip()]

        instructions = [s.strip() for s in request.form.getlist('instruction[]') if s.strip()]

        subs = {}
        for key in ['dairy_free', 'gluten_free', 'vegan', 'vegetarian', 'low_carb']:
            val = request.form.get(f'sub_{key}', '').strip()
            if val:
                subs[key] = val

        recipe = {
            'name':      name,
            'meal_type': meal_type,
            'servings':  servings,
            'prep_time': prep_time,
            'cook_time': cook_time,
            'macros_per_serving': {
                'calories': calories, 'protein_g': protein,
                'carbs_g': carbs, 'fat_g': fat,
            },
            'ingredients':  ingredients,
            'instructions': instructions,
        }
        if subs:
            recipe['substitutions'] = subs
        notes = request.form.get('notes', '').strip()
        if notes:
            recipe['notes'] = notes

        DATA_DIR.mkdir(exist_ok=True)
        slug = re.sub(r"[^a-z0-9]+", '-', name.lower()).strip('-')
        (DATA_DIR / f'{slug}.json').write_text(json.dumps(recipe, indent=2), encoding='utf-8')
        return redirect(url_for('recipes'))

    rv = {'name':'','meal_type':'','servings':4,'prep_time':'','cook_time':'',
          'calories':'','protein':'','carbs':'','fat':'','notes':'',
          'ingredients':[],'instructions':[],
          'sub_dairy_free':'','sub_gluten_free':'','sub_vegan':'','sub_vegetarian':'','sub_low_carb':''}
    return render_template_string(ADD_RECIPE_HTML, meal_types=MEAL_TYPES, rv=rv,
                                  page_title='Add Recipe', edit_mode=False)


@app.route('/recipe/<slug>')
def recipe_detail(slug):
    path = DATA_DIR / f'{slug}.json'
    if not path.exists():
        return redirect(url_for('recipes'))
    recipe = json.load(open(path, encoding='utf-8'))
    wb   = load_wb()
    have = {str(d.get('Item', '')).lower() for _, _, d in get_all_rows(wb) if d.get('Item')}
    used_count     = request.args.get('used_count')
    not_found_count = request.args.get('not_found_count')
    kwargs = {'recipe': recipe, 'slug': slug, 'have': have}
    if used_count is not None:
        kwargs['used_count']      = int(used_count)
        kwargs['not_found_count'] = int(not_found_count or 0)
    return render_template_string(RECIPE_DETAIL_HTML, **kwargs)


@app.route('/recipe/<slug>/use', methods=['POST'])
def recipe_use(slug):
    path = DATA_DIR / f'{slug}.json'
    if not path.exists():
        return redirect(url_for('recipes'))

    recipe = json.load(open(path, encoding='utf-8'))
    wb = load_wb()
    used_count = 0
    not_found_count = 0

    exclusions = load_exclusions()
    excluded_names = {ex.get('name', '').lower() for ex in exclusions}

    for ing in recipe.get('ingredients', []):
        item_name  = (ing.get('item', '') if isinstance(ing, dict) else str(ing)).strip()
        amount_str = (ing.get('amount', '') if isinstance(ing, dict) else '').strip()
        if not item_name:
            continue

        # Skip items on the exclusion list — never auto-decrement these
        ilow = item_name.lower()
        if any(ilow == ex or ilow in ex or ex in ilow for ex in excluded_names):
            continue

        # Parse the leading number from the amount string
        m = re.search(r'[\d.]+', amount_str)
        qty = float(m.group()) if m else 1.0

        # Search all pantry sheets for a fuzzy name match
        found = False
        for sheet_name in [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]:
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                cell = str(ws.cell(row=r, column=1).value or '').strip().lower()
                if not cell:
                    continue
                if cell == item_name.lower() or item_name.lower() in cell or cell in item_name.lower():
                    try:
                        current = float(str(ws.cell(row=r, column=2).value or 0))
                    except (ValueError, TypeError):
                        current = 0
                    new_qty = max(0, current - qty)
                    ws.cell(row=r, column=2).value = int(new_qty) if new_qty == int(new_qty) else new_qty
                    found = True
                    used_count += 1
                    break
            if found:
                break

        if not found:
            not_found_count += 1

    save_wb(wb)
    return redirect(url_for('recipe_detail', slug=slug,
                            used_count=used_count, not_found_count=not_found_count))


@app.route('/recipe/delete/<slug>')
def recipe_delete(slug):
    path = DATA_DIR / f'{slug}.json'
    if path.exists():
        path.unlink()
    return redirect(url_for('recipes'))


@app.route('/recipe/edit/<slug>', methods=['GET', 'POST'])
def recipe_edit(slug):
    path = DATA_DIR / f'{slug}.json'
    if not path.exists():
        return redirect(url_for('recipes'))

    if request.method == 'POST':
        name      = request.form.get('name', '').strip()
        meal_type = request.form.get('meal_type', '')
        servings  = int(request.form.get('servings') or 1)
        prep_time = request.form.get('prep_time', '').strip()
        cook_time = request.form.get('cook_time', '').strip()

        try:
            calories = float(request.form.get('calories') or 0)
            protein  = float(request.form.get('protein') or 0)
            carbs    = float(request.form.get('carbs') or 0)
            fat      = float(request.form.get('fat') or 0)
        except ValueError:
            calories = protein = carbs = fat = 0

        amounts         = request.form.getlist('amount[]')
        ingredients_raw = request.form.getlist('ingredient[]')
        ingredients     = [{'amount': a.strip(), 'item': i.strip()}
                           for a, i in zip(amounts, ingredients_raw) if i.strip()]
        instructions    = [s.strip() for s in request.form.getlist('instruction[]') if s.strip()]

        subs = {}
        for key in ['dairy_free', 'gluten_free', 'vegan', 'vegetarian', 'low_carb']:
            val = request.form.get(f'sub_{key}', '').strip()
            if val:
                subs[key] = val

        recipe = {
            'name': name, 'meal_type': meal_type, 'servings': servings,
            'prep_time': prep_time, 'cook_time': cook_time,
            'macros_per_serving': {'calories': calories, 'protein_g': protein,
                                   'carbs_g': carbs, 'fat_g': fat},
            'ingredients': ingredients, 'instructions': instructions,
        }
        if subs:
            recipe['substitutions'] = subs
        notes = request.form.get('notes', '').strip()
        if notes:
            recipe['notes'] = notes

        # If name changed, rename the file
        new_slug = re.sub(r"[^a-z0-9]+", '-', name.lower()).strip('-')
        if new_slug != slug:
            path.unlink()
        (DATA_DIR / f'{new_slug}.json').write_text(json.dumps(recipe, indent=2), encoding='utf-8')
        return redirect(url_for('recipe_detail', slug=new_slug))

    existing = json.load(open(path, encoding='utf-8'))
    m = existing.get('macros_per_serving', {})
    subs = existing.get('substitutions', {})
    rv = {
        'name':       existing.get('name', ''),
        'meal_type':  existing.get('meal_type', ''),
        'servings':   existing.get('servings', 4),
        'prep_time':  existing.get('prep_time', ''),
        'cook_time':  existing.get('cook_time', ''),
        'calories':   m.get('calories', ''),
        'protein':    m.get('protein_g', ''),
        'carbs':      m.get('carbs_g', ''),
        'fat':        m.get('fat_g', ''),
        'notes':      existing.get('notes', ''),
        'ingredients':   existing.get('ingredients', []),
        'instructions':  existing.get('instructions', []),
        'sub_dairy_free':  subs.get('dairy_free', ''),
        'sub_gluten_free': subs.get('gluten_free', ''),
        'sub_vegan':       subs.get('vegan', ''),
        'sub_vegetarian':  subs.get('vegetarian', ''),
        'sub_low_carb':    subs.get('low_carb', ''),
    }
    return render_template_string(ADD_RECIPE_HTML, meal_types=MEAL_TYPES, rv=rv,
                                  page_title=f'Edit: {rv["name"]}', edit_mode=True)


@app.route('/restock')
def restock():
    return redirect(url_for('shopping'))


@app.route('/shopping')
def shopping():
    wb   = load_wb()
    rows = get_all_rows(wb)
    mins = get_minimums()

    stock = {}
    units = {}
    for _, _, d in rows:
        name = str(d.get('Item') or '').strip().lower()
        if not name:
            continue
        try:
            qty = float(str(d.get('Quantity') or 0))
        except ValueError:
            qty = 0
        stock[name] = stock.get(name, 0) + qty
        if name not in units:
            units[name] = str(d.get('Unit') or '').strip()

    low = []
    for item, minimum, _ in mins:
        key     = item.lower()
        current = stock.get(key, 0)
        if current < minimum:
            low.append((item, current, minimum, units.get(key, '')))

    have = {str(d.get('Item', '')).lower() for _, _, d in rows if d.get('Item')}

    all_recipes = []
    if DATA_DIR.exists():
        for p in sorted(DATA_DIR.glob('*.json')):
            try:
                r = json.load(open(p, encoding='utf-8'))
                r['slug'] = p.stem
                all_recipes.append(r)
            except Exception:
                continue
    all_recipes.sort(key=lambda x: x.get('name', ''))

    selected_recipe = request.args.get('recipe', '')
    missing_ingredients = []
    if selected_recipe:
        chosen = next((r for r in all_recipes if r['slug'] == selected_recipe), None)
        if chosen:
            for ing in chosen.get('ingredients', []):
                item_name = ing.get('item', '').lower()
                if not any(item_name in h or h in item_name for h in have):
                    missing_ingredients.append(ing)

    out_of_stock_exclusions = [ex for ex in load_exclusions() if not ex.get('in_stock', True)]

    return render_template_string(SHOPPING_HTML, low=low, all_recipes=all_recipes,
                                  selected_recipe=selected_recipe,
                                  missing_ingredients=missing_ingredients,
                                  shopping_list=load_shopping_list(),
                                  out_of_stock_exclusions=out_of_stock_exclusions)


@app.route('/shopping/list/add-recipe', methods=['POST'])
def shopping_list_add_recipe():
    slug = request.form.get('recipe', '')
    if not slug or not DATA_DIR.exists():
        return redirect(url_for('shopping', recipe=slug))
    path = DATA_DIR / f'{slug}.json'
    if not path.exists():
        return redirect(url_for('shopping', recipe=slug))

    recipe = json.loads(path.read_text(encoding='utf-8'))
    recipe_name = recipe.get('name', slug)

    wb   = load_wb()
    have = {str(d.get('Item', '')).lower() for _, _, d in get_all_rows(wb) if d.get('Item')}

    items = load_shopping_list()
    existing_names = {i['name'].lower() for i in items}

    exclusions = load_exclusions()
    excluded_names = {ex.get('name', '').lower() for ex in exclusions}

    for ing in recipe.get('ingredients', []):
        item_name = ing.get('item', '').strip()
        if not item_name:
            continue
        item_lower = item_name.lower()
        if any(item_lower == ex or item_lower in ex or ex in item_lower for ex in excluded_names):
            continue
        if any(item_lower in h or h in item_lower for h in have):
            continue
        if item_lower in existing_names:
            continue
        items.append({
            'id':      str(int(datetime.now().timestamp() * 1000)) + str(len(items)),
            'name':    item_name,
            'amount':  ing.get('amount', ''),
            'note':    f'For: {recipe_name}',
            'checked': False,
        })
        existing_names.add(item_lower)

    save_shopping_list(items)
    return redirect(url_for('shopping', recipe=slug))


@app.route('/shopping/list/toggle/<item_id>', methods=['POST'])
def shopping_list_toggle(item_id):
    items = load_shopping_list()
    for item in items:
        if item['id'] == item_id:
            item['checked'] = not item['checked']
            break
    save_shopping_list(items)
    recipe = request.args.get('recipe', '')
    return redirect(url_for('shopping', recipe=recipe) if recipe else url_for('shopping'))


@app.route('/shopping/list/remove/<item_id>')
def shopping_list_remove(item_id):
    items = [i for i in load_shopping_list() if i['id'] != item_id]
    save_shopping_list(items)
    return redirect(url_for('shopping'))


@app.route('/shopping/list/clear-all', methods=['POST'])
def shopping_list_clear_all():
    save_shopping_list([])
    return redirect(url_for('shopping'))


@app.route('/shopping/list/clear-checked', methods=['POST'])
def shopping_list_clear_checked():
    items = [i for i in load_shopping_list() if not i['checked']]
    save_shopping_list(items)
    return redirect(url_for('shopping'))


@app.route('/add/bulk', methods=['GET', 'POST'])
def bulk_add():
    location_sheets = get_location_sheets()
    results = []
    if request.method == 'POST':
        bulk_text        = request.form.get('bulk', '')
        default_location = request.form.get('default_location', location_sheets[0] if location_sheets else '')
        wb = load_wb()

        for line in bulk_text.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(',')]
            item     = to_title(parts[0]) if parts else ''
            quantity = parts[1] if len(parts) > 1 else None
            unit     = parts[2] if len(parts) > 2 else None
            location = parts[3] if len(parts) > 3 else default_location
            if location not in location_sheets:
                location = default_location

            if not item:
                continue

            ws = wb[location]
            unit_lower = (unit or '').lower()
            existing = next((r for r in range(2, ws.max_row + 1)
                             if str(ws.cell(row=r, column=1).value or '').lower() == item.lower()
                             and str(ws.cell(row=r, column=3).value or '').lower() == unit_lower), None)
            if existing:
                try:
                    cur = float(str(ws.cell(row=existing, column=2).value or 0))
                    add_qty = float(quantity) if quantity else 1
                    merged = cur + add_qty
                    ws.cell(row=existing, column=2).value = int(merged) if merged == int(merged) else merged
                    results.append(f'✅ Updated <strong>{item}</strong> (merged)')
                except (ValueError, TypeError):
                    results.append(f'⚠️ Could not merge quantities for <strong>{item}</strong>')
            else:
                row = next((r for r in range(2, ws.max_row + 2) if not ws.cell(row=r, column=1).value),
                           ws.max_row + 1)
                ws.cell(row=row, column=1).value = item
                ws.cell(row=row, column=2).value = quantity
                ws.cell(row=row, column=3).value = unit
                ws.cell(row=row, column=4).value = location
                results.append(f'✅ Added <strong>{item}</strong> → {location}')

        save_wb(wb)
        wb2     = load_wb()
        all_raw = get_all_rows(wb2)
        rows    = [(f'{s}:{r}', d) for s, r, d in all_raw]
        return render_template_string(INDEX_HTML, items=rows, total=len(rows),
                                      locations=location_sheets, q='', loc='',
                                      tab='add', bulk_results=results, units=load_units())

    return redirect(url_for('index', tab='add'))


@app.route('/settings')
def settings():
    env = read_env()
    token             = env.get('DISCORD_TOKEN', '')
    alert_channel     = env.get('DISCORD_ALERT_CHANNEL', '')
    backup_mode       = env.get('BACKUP_MODE', 'actions')
    backup_every_n    = int(env.get('BACKUP_EVERY_N', '10'))
    backup_interval   = float(env.get('BACKUP_INTERVAL_HRS', '24'))
    backup_max        = int(env.get('BACKUP_MAX', '10'))
    backup_last_label, backup_count = get_backup_info()
    active_tab        = request.args.get('tab', 'locations')
    return render_template_string(SETTINGS_HTML, token_set=bool(token),
                                  alert_channel=alert_channel,
                                  minimums=get_minimums(),
                                  locations=get_location_info(),
                                  backups=get_backups(),
                                  backup_mode=backup_mode,
                                  backup_every_n=backup_every_n,
                                  backup_interval=backup_interval,
                                  backup_max=backup_max,
                                  backup_last_label=backup_last_label,
                                  backup_count=backup_count,
                                  exclusions=load_exclusions(),
                                  units=get_unit_info(),
                                  active_tab=active_tab)


@app.route('/locations/add', methods=['POST'])
def locations_add():
    name = to_title(request.form.get('name', '').strip())
    if name and name not in EXCLUDE_SHEETS:
        existing = get_location_sheets()
        if name not in existing:
            wb = load_wb()
            ws = wb.create_sheet(name)
            for col, header in enumerate(COLS, 1):
                ws.cell(row=1, column=col).value = header
            wb.save(XLSX)
    return redirect(url_for('settings', tab='locations'))


@app.route('/locations/delete/<name>')
def locations_delete(name):
    if name in EXCLUDE_SHEETS:
        return redirect(url_for('settings', tab='locations'))
    wb = load_wb()
    if name in wb.sheetnames:
        ws = wb[name]
        count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value)
        if count == 0:
            del wb[name]
            wb.save(XLSX)
    return redirect(url_for('settings', tab='locations'))


@app.route('/exclusions/add', methods=['POST'])
def exclusions_add():
    name      = to_title(request.form.get('name', '').strip())
    container = request.form.get('container', '').strip()
    if not name:
        return redirect(url_for('settings', tab='exclusions'))
    items = load_exclusions()
    if not any(ex.get('name', '').lower() == name.lower() for ex in items):
        items.append({'name': name, 'container': container, 'in_stock': True})
        save_exclusions(items)
    return redirect(url_for('settings', tab='exclusions'))


@app.route('/exclusions/delete/<path:name>')
def exclusions_delete(name):
    items = [ex for ex in load_exclusions() if ex.get('name', '').lower() != name.lower()]
    save_exclusions(items)
    return redirect(url_for('settings', tab='exclusions'))


@app.route('/exclusions/toggle/<path:name>', methods=['POST'])
def exclusions_toggle(name):
    items = load_exclusions()
    for ex in items:
        if ex.get('name', '').lower() == name.lower():
            ex['in_stock'] = not ex.get('in_stock', True)
            break
    save_exclusions(items)
    referer = request.referrer or url_for('settings', tab='exclusions')
    return redirect(referer)


@app.route('/units/add', methods=['POST'])
def units_add():
    unit = request.form.get('unit', '').strip()
    if unit:
        units = load_units()
        if unit.lower() not in [u.lower() for u in units]:
            units.append(unit)
            save_units(units)
    return redirect(url_for('settings', tab='units'))


@app.route('/units/delete/<unit>')
def units_delete(unit):
    _, count = next(((u, c) for u, c in get_unit_info() if u.lower() == unit.lower()), (None, 0))
    if count == 0:
        units = [u for u in load_units() if u.lower() != unit.lower()]
        save_units(units)
    return redirect(url_for('settings', tab='units'))


@app.route('/units/reorder', methods=['POST'])
def units_reorder():
    data = request.get_json(silent=True) or {}
    new_order = data.get('order', [])
    current = load_units()
    current_lower = [u.lower() for u in current]
    ordered = [u for u in new_order if u.lower() in current_lower]
    rest = [u for u in current if u.lower() not in [x.lower() for x in ordered]]
    save_units(ordered + rest)
    return ('', 204)


@app.route('/locations/reorder', methods=['POST'])
def locations_reorder():
    data = request.get_json(silent=True) or {}
    new_order = data.get('order', [])
    wb = load_wb()
    valid = [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]
    new_order = [n for n in new_order if n in valid]
    excluded = [s for s in wb.sheetnames if s in EXCLUDE_SHEETS]
    full_order = new_order + excluded
    for target_idx, name in enumerate(full_order):
        cur_idx = wb.sheetnames.index(name)
        if cur_idx != target_idx:
            wb.move_sheet(name, offset=target_idx - cur_idx)
    wb.save(XLSX)
    return ('', 204)


@app.route('/settings/discord', methods=['POST'])
def settings_discord():
    action = request.form.get('action', 'save')
    token  = request.form.get('token', '').strip()

    alert_channel = request.form.get('alert_channel', '').strip()
    env = read_env()

    if token:
        env['DISCORD_TOKEN'] = token

    env['DISCORD_ALERT_CHANNEL'] = alert_channel
    write_env(env)

    if action == 'stop':
        stop_bot()
    elif action in ('save_restart', 'restart'):
        start_bot()
    elif action == 'remove':
        stop_bot()
        env.pop('DISCORD_TOKEN', None)
        write_env(env)

    return redirect(url_for('settings', tab='discord'))


@app.route('/settings/bot-status')
def settings_bot_status():
    return jsonify({'running': bot_running()})


@app.route('/settings/bot-test', methods=['POST'])
def settings_bot_test():
    import requests as req
    token = request.form.get('token', '').strip()
    if not token:
        return jsonify({'ok': False, 'msg': 'No token provided.'})
    try:
        r = req.get('https://discord.com/api/v10/users/@me',
                    headers={'Authorization': f'Bot {token}'}, timeout=8)
        if r.status_code == 200:
            name = r.json().get('username', '?')
            return jsonify({'ok': True, 'msg': f'✅ Token valid — bot is "{name}"'})
        elif r.status_code == 401:
            return jsonify({'ok': False, 'msg': '❌ Invalid token — Discord rejected it (401). Reset the token in the developer portal and try again.'})
        else:
            return jsonify({'ok': False, 'msg': f'❌ Discord returned {r.status_code}.'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'❌ Could not reach Discord: {e}'})


@app.route('/settings/bot-log')
def settings_bot_log():
    if not BOT_LOG_FILE.exists():
        return jsonify({'log': ''})
    text  = BOT_LOG_FILE.read_text(encoding='utf-8', errors='replace')
    # Only show the most recent session (after the last separator line)
    marker = '--- Bot started'
    idx = text.rfind(marker)
    session = text[idx:] if idx != -1 else text
    return jsonify({'log': session.strip()})


@app.route('/settings/appearance', methods=['POST'])
def settings_appearance():
    color = request.form.get('accent_color', '#6B2D8B').strip()
    if re.match(r'^#[0-9A-Fa-f]{6}$', color):
        env = read_env()
        env['ACCENT_COLOR'] = color
        write_env(env)
    return redirect(url_for('settings', tab='appearance'))


@app.route('/settings/backups', methods=['POST'])
def settings_backups():
    env  = read_env()
    mode = request.form.get('backup_mode', 'actions')
    if mode not in ('actions', 'interval', 'manual'):
        mode = 'actions'
    try:
        every_n = max(1, min(500, int(request.form.get('backup_every_n', '10'))))
    except ValueError:
        every_n = 10
    try:
        interval_hrs = max(0.25, float(request.form.get('backup_interval_hrs', '24')))
    except ValueError:
        interval_hrs = 24
    try:
        max_backups = max(1, min(50, int(request.form.get('max_backups', '10'))))
    except ValueError:
        max_backups = 10
    env['BACKUP_MODE']         = mode
    env['BACKUP_EVERY_N']      = str(every_n)
    env['BACKUP_INTERVAL_HRS'] = str(interval_hrs)
    env['BACKUP_MAX']          = str(max_backups)
    write_env(env)
    return redirect(url_for('settings', tab='backups'))


@app.route('/backups/now', methods=['POST'])
def backups_now():
    backup_wb(force=True)
    return redirect(url_for('settings', tab='backups'))


@app.route('/backups/download/<filename>')
def backups_download(filename):
    path = BACKUP_DIR / Path(filename).name
    if path.exists() and path.parent.resolve() == BACKUP_DIR.resolve():
        return send_file(path, as_attachment=True)
    return redirect(url_for('settings', tab='backups'))


@app.route('/backups/restore/<filename>', methods=['POST'])
def backups_restore(filename):
    path = BACKUP_DIR / Path(filename).name
    if path.exists() and path.parent.resolve() == BACKUP_DIR.resolve():
        backup_wb(force=True)
        if path.suffix.lower() == '.zip':
            with zipfile.ZipFile(path, 'r') as zf:
                names = zf.namelist()
                if XLSX.name in names:
                    zf.extract(XLSX.name, XLSX.parent)
                if UNITS_FILE.name in names:
                    zf.extract(UNITS_FILE.name, UNITS_FILE.parent)
                if EXCLUSIONS_FILE.name in names:
                    zf.extract(EXCLUSIONS_FILE.name, EXCLUSIONS_FILE.parent)
                DATA_DIR.mkdir(exist_ok=True)
                for name in names:
                    if name.startswith('data/') and name.endswith('.json'):
                        zf.extract(name, Path(__file__).parent)
        else:
            shutil.copy2(path, XLSX)
    return redirect(url_for('settings', tab='backups'))


@app.route('/backups/upload', methods=['POST'])
def backups_upload():
    f = request.files.get('backup_file')
    if not f:
        return redirect(url_for('settings', tab='backups'))
    fname = f.filename.lower()
    if not (fname.endswith('.zip') or fname.endswith('.xlsx')):
        return redirect(url_for('settings', tab='backups'))
    BACKUP_DIR.mkdir(exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    ext  = '.zip' if fname.endswith('.zip') else '.xlsx'
    dest = BACKUP_DIR / f'pantry_backup_upload_{ts}{ext}'
    f.save(str(dest))
    env = read_env()
    max_backups = max(1, int(env.get('BACKUP_MAX', '10')))
    all_files = sorted(
        list(BACKUP_DIR.glob('*.zip')) + list(BACKUP_DIR.glob('*.xlsx')),
        key=lambda p: p.stat().st_mtime
    )
    for old in all_files[:-max_backups]:
        old.unlink(missing_ok=True)
    return redirect(url_for('settings', tab='backups'))


@app.route('/recipes/download-pdf')
def download_pdf():
    try:
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).parent))
        from generate_pdf import generate_binder_bytes
    except ImportError:
        return 'reportlab is not installed. Run: pip install reportlab', 500

    data = generate_binder_bytes()
    if not data:
        return 'No recipes found.', 404

    from datetime import datetime
    filename = f'recipe_binder_{datetime.now().strftime("%Y%m%d")}.pdf'
    return send_file(
        io.BytesIO(data),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    _kill_stale_bot()
    print('Starting Pantri at http://localhost:5000')

    def _port_open(port):
        try:
            with socket.create_connection(('localhost', port), timeout=0.5):
                return True
        except OSError:
            return False

    if not _port_open(5000):
        threading.Timer(1.2, webbrowser.open, args=['http://localhost:5000']).start()

    start_bot()
    app.run(host='0.0.0.0', port=5000)
